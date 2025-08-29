# views.py
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework import status
from collections import Counter,defaultdict
from django.db.models import Count, Q, Case, When, F,Value ,CharField, Max, ExpressionWrapper, DecimalField
from django.db.models.functions import TruncMonth, ExtractMonth, ExtractYear
from .models import Audit_ip_fh_model, AuditCSG, AuditATR, CustomAudit,AuditDoublonsIP
from .serializers import FileUploadSerializer,CustomAuditSerializer,AuditDoublonsIPSerializer
from django.db import models
from django.http import JsonResponse
from django.utils import timezone
import openpyxl
from dateutil.relativedelta import relativedelta
from django.http import HttpResponse
from datetime import date, timedelta, datetime,time
from rest_framework.parsers import MultiPartParser, FormParser
import pandas as pd
from rest_framework import serializers


########################################### Audit IP FH ##################################################

class FileUploadSerializer(serializers.Serializer):
    file = serializers.FileField()

class ExcelUploadView(APIView):
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, *args, **kwargs):
        serializer = FileUploadSerializer(data=request.data)
        if serializer.is_valid():
            excel_files = request.FILES.getlist('file')
            for excel_file in excel_files:
                try:
                    self.process_excel(excel_file)
                except Exception as e:
                    return Response({'message': f'Error processing Excel file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
            return Response({'message': 'Excel file processed successfully'}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def process_excel(self, file):
      df = pd.read_excel(file)
      for _, row in df.iterrows():

        date_demande_swap_dci = pd.to_datetime(row['date_demande_swap_dci'], errors='coerce')
        if pd.isna(date_demande_swap_dci):
            date_demande_swap_dci = None

        data_insertion_date = pd.to_datetime(row['Data_Insertion_Date'], errors='coerce')
        if pd.isna(data_insertion_date):
            data_insertion_date = None
        else:
            data_insertion_date = data_insertion_date.date()

        Audit_ip_fh_model.objects.create(
            region=row['region'],
            id_lien=row['id_lien'],
            equipement=row['equipement'],
            site_geo=row['site_geo'],
            site_theo=row['site_theo'],
            region_1=row['region_1'],
            installe=row['installe'],
            type_extremite=row['type_extremite'],
            adresse_ip=row['adresse_ip'] if not pd.isna(row['adresse_ip']) else None,
            statut_ip=row['statut_ip'],
            constructeur=row['constructeur'],
            materiel=row['materiel'],
            etat=row['etat'],
            nature=row['nature'],
            statut=row['statut'],
            statut_lp35=row['statut_lp35'],
            statut_swap_fh=row['statut_swap_fh'],
            ip_omc=row['ip_omc'] if not pd.isna(row['ip_omc']) else None,
            tdj_osa=row['tdj_osa'],
            check_ip=row['check_ip'],
            tdj_cdr=row['tdj_cdr'],
            tdj_swap_ip=row['tdj_swap_ip'],
            date_demande_swap_dci=date_demande_swap_dci,
            nb_fht=row['nb_fht'],
            statut_coherence_adresses_ip=row['statut_coherence_adresses_ip'],
            porteur=row['porteur'],
            action_description=row['action_description'],
            type_incoherence=row['type_incoherence'] if not pd.isna(row['type_incoherence']) else '',
            Data_Insertion_Date=data_insertion_date
        )

class AuditIpFhIncoherenceTdjCdrNew(APIView):
    def get(self, request):
        total_records_with_tdj_cdr = Audit_ip_fh_model.objects.filter(tdj_cdr='TDJ CDR').count()

        latest_insertion_date = Audit_ip_fh_model.objects.filter(
            tdj_cdr='TDJ CDR'
        ).aggregate(latest_date=Max('Data_Insertion_Date'))['latest_date']

        if latest_insertion_date:
            month_start = latest_insertion_date.replace(day=1)
            if latest_insertion_date.month == 12:
                month_end = latest_insertion_date.replace(year=latest_insertion_date.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                month_end = latest_insertion_date.replace(month=latest_insertion_date.month + 1, day=1) - timedelta(days=1)

            ip_coherent_bdr_bde = Audit_ip_fh_model.objects.filter(
                Q(type_incoherence='Adresse IP Existante dans BDR et BDE et conformes'),
                tdj_cdr='TDJ CDR',
                Data_Insertion_Date__range=(month_start, month_end)
            ).count()

            ip_incoherent_bdr_bde = Audit_ip_fh_model.objects.filter(
                type_incoherence='Adresse IP Existante dans BDR et BDE mais non conformes',
                tdj_cdr='TDJ CDR',
                Data_Insertion_Date__range=(month_start, month_end)
            ).count()

            ip_bdr_uniquement = Audit_ip_fh_model.objects.filter(
                type_incoherence='Adresse IP Existante dans BDR uniquement',
                tdj_cdr='TDJ CDR',
                Data_Insertion_Date__range=(month_start, month_end)
            ).count()

            ip_bde_uniquement = Audit_ip_fh_model.objects.filter(         
                type_incoherence='Adresse IP Existante dans BDE uniquement',
                tdj_cdr='TDJ CDR',
                Data_Insertion_Date__range=(month_start, month_end)
            ).count()

            fht_incoherence_swap_count = Audit_ip_fh_model.objects.filter(
                action_description='SWAP IP Prévu par la PROD',
                tdj_cdr='TDJ CDR',
                Data_Insertion_Date__range=(month_start, month_end)
            ).count()

            fht_incoherence_adeployer_count = Audit_ip_fh_model.objects.filter(
                action_description='FHT à déployer',
                tdj_cdr='TDJ CDR',
                Data_Insertion_Date__range=(month_start, month_end)
            ).count() 

        else:
            ip_coherent_bdr_bde = 0
            ip_incoherent_bdr_bde = 0
            ip_bdr_uniquement = 0
            ip_bde_uniquement = 0
            fht_incoherence_swap_count = 0
            fht_incoherence_adeployer_count = 0

        percentage_ip_coherent = (ip_coherent_bdr_bde / total_records_with_tdj_cdr) * 100 if total_records_with_tdj_cdr > 0 else 0.0
        percentage_ip_incoherent = (ip_incoherent_bdr_bde / total_records_with_tdj_cdr) * 100 if total_records_with_tdj_cdr > 0 else 0.0
        percentage_ip_bde_uniquement = (ip_bde_uniquement / total_records_with_tdj_cdr) * 100 if total_records_with_tdj_cdr > 0 else 0.0
        percentage_ip_bdr_uniquement = (ip_bdr_uniquement / total_records_with_tdj_cdr) * 100 if total_records_with_tdj_cdr > 0 else 0.0
        percentage_incoherence_swap = (fht_incoherence_swap_count / total_records_with_tdj_cdr) * 100 if total_records_with_tdj_cdr > 0 else 0.0
        percentage_incoherence_fhtadeployer = (fht_incoherence_adeployer_count / total_records_with_tdj_cdr) * 100 if total_records_with_tdj_cdr > 0 else 0.0

        percentage_ip_coherent = round(percentage_ip_coherent)
        percentage_ip_incoherent = round(percentage_ip_incoherent)
        percentage_ip_bde_uniquement = round(percentage_ip_bde_uniquement)
        percentage_ip_bdr_uniquement = round(percentage_ip_bdr_uniquement)
        percentage_incoherence_swap = round(percentage_incoherence_swap)
        percentage_incoherence_fhtadeployer = round(percentage_incoherence_fhtadeployer)

        result_data = {
            'total_records_with_tdj_cdr': total_records_with_tdj_cdr,
            'percentage_ip_coherent': percentage_ip_coherent,
            'percentage_ip_incoherent': percentage_ip_incoherent,
            'percentage_ip_bde_uniquement': percentage_ip_bde_uniquement,
            'percentage_ip_bdr_uniquement': percentage_ip_bdr_uniquement,
            'percentage_incoherence_swap': percentage_incoherence_swap,
            'total_records': total_records_with_tdj_cdr,
            'ip_bdr_uniquement': ip_bdr_uniquement,
            'ip_bde_uniquement': ip_bde_uniquement,
            'ip_coherent_bdr_bde': ip_coherent_bdr_bde,
            'ip_incoherent_bdr_bde': ip_incoherent_bdr_bde,
            'fht_incoherence_swap_count': fht_incoherence_swap_count,
            'percentage_incoherence_fhtadeployer': percentage_incoherence_fhtadeployer,
            'fht_incoherence_adeployer_count': fht_incoherence_adeployer_count,
        }

        return JsonResponse(result_data)

class AuditIpFhTauxDeCoherence(APIView):
    def calculate_nok_percentage(self):
        # Get the latest insertion date
        latest_insertion_date = Audit_ip_fh_model.objects.filter(
            tdj_cdr='TDJ CDR'
        ).aggregate(latest_date=Max('Data_Insertion_Date'))['latest_date']

        if latest_insertion_date:
            # Calculate the start and end dates of the month for the latest insertion date
            month_start = latest_insertion_date.replace(day=1)
            if latest_insertion_date.month == 12:
                month_end = latest_insertion_date.replace(year=latest_insertion_date.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                month_end = latest_insertion_date.replace(month=latest_insertion_date.month + 1, day=1) - timedelta(days=1)

            total_records_with_tdj_cdr = Audit_ip_fh_model.objects.filter(
                tdj_cdr='TDJ CDR',
                Data_Insertion_Date__range=(month_start, month_end)
            ).count()

            nbr_fht_coherent = Audit_ip_fh_model.objects.filter(
                Q(statut_coherence_adresses_ip='OK') | Q(statut_coherence_adresses_ip='NOK Justifié'),
                tdj_cdr='TDJ CDR',
                Data_Insertion_Date__range=(month_start, month_end)
            ).count()

            nbr_fht_incoherent = total_records_with_tdj_cdr - nbr_fht_coherent

            if total_records_with_tdj_cdr == 0:
                return {
                    'FHT_Coherent': 0,
                    'FHT_Incoherent': 0
                }

            percentage_fht_coherent = (nbr_fht_coherent / total_records_with_tdj_cdr) * 100
            percentage_fht_incoherent = (nbr_fht_incoherent / total_records_with_tdj_cdr) * 100

            return {
                'FHT_Coherent': round(percentage_fht_coherent),
                'FHT_Incoherent': round(percentage_fht_incoherent),
            }
        else:
            return {
                'FHT_Coherent': 0,
                'FHT_Incoherent': 0
            }

    def get(self, request, *args, **kwargs):
        nok_percentages = self.calculate_nok_percentage()
        return JsonResponse(nok_percentages)
    
class AuditIpVolumeTdjCdr(APIView):
    def calculate_nok_volume(self):
        # Get the latest insertion date
        latest_insertion_date = Audit_ip_fh_model.objects.filter(
            tdj_cdr='TDJ CDR'
        ).aggregate(latest_date=Max('Data_Insertion_Date'))['latest_date']

        if latest_insertion_date:
            # Calculate the start and end dates of the month for the latest insertion date
            month_start = latest_insertion_date.replace(day=1)
            if latest_insertion_date.month == 12:
                month_end = latest_insertion_date.replace(year=latest_insertion_date.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                month_end = latest_insertion_date.replace(month=latest_insertion_date.month + 1, day=1) - timedelta(days=1)

            total_records_with_tdj_cdr = Audit_ip_fh_model.objects.filter(
                tdj_cdr='TDJ CDR',
                Data_Insertion_Date__range=(month_start, month_end)
            ).count()

            nbr_fht_coherent = Audit_ip_fh_model.objects.filter(
                Q(statut_coherence_adresses_ip='OK') | Q(statut_coherence_adresses_ip='NOK Justifié'),     
                tdj_cdr='TDJ CDR',
                Data_Insertion_Date__range=(month_start, month_end)
            ).count()

            nbr_fht_incoherent = total_records_with_tdj_cdr - nbr_fht_coherent

            if total_records_with_tdj_cdr == 0:
                return {
                    'FHT_Coherent': 0,
                    'FHT_Incoherent': 0
                }

            return {
                'Nbr_FHT_Coherent': nbr_fht_coherent,
                'Nbr_FHT_Incoherent': nbr_fht_incoherent,
            }
        else:
            return {
                'Nbr_FHT_Coherent': 0,
                'Nbr_FHT_Incoherent': 0
            }

    def get(self, request, *args, **kwargs):
        nok_volume = self.calculate_nok_volume()
        return JsonResponse(nok_volume)
    
class AuditIpFhEvolutionTauxDeCoherence(APIView):
    def calculate_nok_percentage(self):
        # Get the latest and second latest insertion dates
        latest_injection_date = Audit_ip_fh_model.objects.filter(
            tdj_cdr='TDJ CDR'
        ).aggregate(latest_date=Max('Data_Insertion_Date'))['latest_date']

        second_latest_injection_date = Audit_ip_fh_model.objects.filter(
            tdj_cdr='TDJ CDR',
            Data_Insertion_Date__lt=latest_injection_date
        ).aggregate(second_latest_date=Max('Data_Insertion_Date'))['second_latest_date']

        if latest_injection_date and second_latest_injection_date:
            latest_month_start = latest_injection_date.replace(day=1)
            if latest_injection_date.month == 12:
                latest_month_end = latest_injection_date.replace(year=latest_injection_date.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                latest_month_end = latest_injection_date.replace(month=latest_injection_date.month + 1, day=1) - timedelta(days=1)

            # Calculate the start and end dates for the second latest month
            second_latest_month_start = second_latest_injection_date.replace(day=1)
            if second_latest_injection_date.month == 12:
                second_latest_month_end = second_latest_injection_date.replace(year=second_latest_injection_date.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                second_latest_month_end = second_latest_injection_date.replace(month=second_latest_injection_date.month + 1, day=1) - timedelta(days=1)

            # Calculate total records and coherent FHT for the latest month
            total_records_with_tdj_cdr_latest = Audit_ip_fh_model.objects.filter(
                tdj_cdr='TDJ CDR',
                Data_Insertion_Date__range=(latest_month_start, latest_month_end)
            ).count()

            nbr_fht_coherent_latest = Audit_ip_fh_model.objects.filter(
                Q(statut_coherence_adresses_ip='OK') | Q(statut_coherence_adresses_ip='NOK Justifié'),
                tdj_cdr='TDJ CDR',
                Data_Insertion_Date__range=(latest_month_start, latest_month_end)
            ).count()

            # Calculate total records and coherent FHT for the second latest month
            total_records_with_tdj_cdr_second_latest = Audit_ip_fh_model.objects.filter(
                tdj_cdr='TDJ CDR',
                Data_Insertion_Date__range=(second_latest_month_start, second_latest_month_end)
            ).count()

            nbr_fht_coherent_second_latest = Audit_ip_fh_model.objects.filter(
                Q(statut_coherence_adresses_ip='OK') | Q(statut_coherence_adresses_ip='NOK Justifié'),
                tdj_cdr='TDJ CDR',
                Data_Insertion_Date__range=(second_latest_month_start, second_latest_month_end)
            ).count()

            # Calculate incoherent FHT counts for both months
            nbr_fht_incoherent_latest = total_records_with_tdj_cdr_latest - nbr_fht_coherent_latest
            nbr_fht_incoherent_second_latest = total_records_with_tdj_cdr_second_latest - nbr_fht_coherent_second_latest

            if total_records_with_tdj_cdr_latest == 0 or total_records_with_tdj_cdr_second_latest == 0:
                return {
                    'FHT_Coherent_Current': 0,
                    'FHT_Incoherent_Current': 0,
                    'FHT_Coherent_Last': 0,
                    'FHT_Incoherent_Last': 0,
                    'Difference': 0,
                    'Positive_Variation': False
                }

            # Calculate percentages for both months
            percentage_fht_coherent_latest = (nbr_fht_coherent_latest / total_records_with_tdj_cdr_latest) * 100
            percentage_fht_incoherent_latest = (nbr_fht_incoherent_latest / total_records_with_tdj_cdr_latest) * 100

            percentage_fht_coherent_second_latest = (nbr_fht_coherent_second_latest / total_records_with_tdj_cdr_second_latest) * 100
            percentage_fht_incoherent_second_latest = (nbr_fht_incoherent_second_latest / total_records_with_tdj_cdr_second_latest) * 100

            # Calculate the difference and positive variation
            difference = percentage_fht_coherent_latest - percentage_fht_coherent_second_latest
            positive_variation = difference > 0.00

            coherent_category = self.calculate_category(percentage_fht_coherent_latest)

            return {
                'FHT_Coherent': round(percentage_fht_coherent_latest),
                'FHT_Incoherent': round(percentage_fht_incoherent_latest),
                'FHT_Coherent_LastMonth': round(percentage_fht_coherent_second_latest),
                'FHT_Incoherent_LastMonth': round(percentage_fht_incoherent_second_latest),
                'Difference': round(abs(difference)),
                'Positive_Variation': positive_variation,
                'Coherent_Category': coherent_category
            }
        else:
            return {
                'FHT_Coherent': 0,
                'FHT_Incoherent': 0,
                'FHT_Coherent_LastMonth': 0,
                'FHT_Incoherent_LastMonth': 0,
                'Difference': 0,
                'Positive_Variation': False,
                'Coherent_Category': 'low'
            }

    def calculate_category(self, percentage):
        if percentage < 60:
            return 'low'
        elif 60 <= percentage <= 90:
            return 'medium'
        else:
            return 'high'

    def get(self, request, *args, **kwargs):
        nok_percentages = self.calculate_nok_percentage()
        return JsonResponse(nok_percentages)    

class ExportAuditIpFhToExcelView(APIView):
    def get(self, request, *args, **kwargs):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="audit_records.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active

        columns = [f.name for f in Audit_ip_fh_model._meta.fields]
        for col_num, column_title in enumerate(columns, start=1):
            ws.cell(row=1, column=col_num, value=column_title)

        rows = Audit_ip_fh_model.objects.all()
        for row_num, row in enumerate(rows, start=2):
            row_data = [getattr(row, field) for field in columns]
            for col_num, cell_value in enumerate(row_data, start=1):
                ws.cell(row=row_num, column=col_num, value=cell_value)

        wb.save(response)
        return response

class GetLatestDataInsertionDateAuditIpFhView(APIView):
    def get(self, request, *args, **kwargs):
        try:
            latest_record = Audit_ip_fh_model.objects.latest('Data_Insertion_Date')
            latest_date = latest_record.Data_Insertion_Date.strftime('%Y-%m-%d')  # Format the date as needed
            response_data = {'latest_date': latest_date}
            return JsonResponse(response_data)
        except Audit_ip_fh_model.DoesNotExist:
            response_data = {'error': 'No records found in the model.'}
            return JsonResponse(response_data, status=404)

class AuditIpFhIncoherenceTdjCdrendetailsBdr(APIView):
    def get(self, request):
        # Count the total number of rows
          # Calculate the date range for the current month
        today = date.today()
        current_month_start = today.replace(day=1)
        current_month_end = today

        # Count the total number of rows with tdj_cdr='TDJ CDR' for the current month
        total_records_with_tdj_cdr = Audit_ip_fh_model.objects.filter(
            tdj_cdr='TDJ CDR',
            Data_Insertion_Date__range=(current_month_start, current_month_end)
        ).count()

        # Count the number of rows for 'Mise à jour de la BDR', 'Nettoyage BDR : Doublon IP à nettoyer', 'NC SWAP FH - BDR à mettre à jour'
        # with tdj_cdr='TDJ CDR' for the current month
        fht_incoherence_maj_bdr_count = Audit_ip_fh_model.objects.filter(
            Q(action_description='Mise à jour de la BDR'),
            tdj_cdr='TDJ CDR',
            Data_Insertion_Date__range=(current_month_start, current_month_end)
        ).count()
        
        fht_incoherence_doublons_ip_bdr_count = Audit_ip_fh_model.objects.filter(
            Q(action_description='Nettoyage BDR : Doublon IP à nettoyer'),
            tdj_cdr='TDJ CDR',
            Data_Insertion_Date__range=(current_month_start, current_month_end)
        ).count()
        
        fht_incoherence_nc_swap_fh_bdr_count = Audit_ip_fh_model.objects.filter(
            Q(action_description='NC SWAP FH - BDR à mettre à jour'),
            tdj_cdr='TDJ CDR',
            Data_Insertion_Date__range=(current_month_start, current_month_end)
        ).count()

        # Count the number of rows for 'NC LP35 - Nettoyage DBE à Faire' with tdj_cdr='TDJ CDR' for the current month
        fht_incoherence_adeployer_count = Audit_ip_fh_model.objects.filter(
            action_description='FHT à déployer',
            tdj_cdr='TDJ CDR',
            Data_Insertion_Date__range=(current_month_start, current_month_end)
        ).count()

        fht_incoherence_bde_count = Audit_ip_fh_model.objects.filter(
            action_description='NC LP35 - Nettoyage DBE à Faire',
            tdj_cdr='TDJ CDR',
            Data_Insertion_Date__range=(current_month_start, current_month_end)
        ).count()

        fht_incoherence_swap_count = Audit_ip_fh_model.objects.filter(
            action_description='SWAP IP Prévu par la PROD',
            tdj_cdr='TDJ CDR',
            Data_Insertion_Date__range=(current_month_start, current_month_end)
        ).count()

        # Count the number of rows for 'Vérification du trafic et mise à jour de BDR ou BDE',
        # 'Adresse IP incohérente entre BDR & BDE' with tdj_cdr='TDJ CDR' for the current month
        fht_incoherence_bdr_ou_bde_verification_trafic_count = Audit_ip_fh_model.objects.filter(
            Q(action_description='Vérification du trafic et mise à jour de BDR ou BDE'),
            tdj_cdr='TDJ CDR',
            Data_Insertion_Date__range=(current_month_start, current_month_end)
        ).count()
        
        fht_incoherence_bdr_ou_bde_ip_incoherence_count = Audit_ip_fh_model.objects.filter(
            Q(action_description='Adresse IP incohérente entre BDR & BDE'),
            tdj_cdr='TDJ CDR',
            Data_Insertion_Date__range=(current_month_start, current_month_end)
        ).count()

        # Calculate the percentage
        percentage_incoherence_maj_bdr_count = (fht_incoherence_maj_bdr_count / total_records_with_tdj_cdr) * 100 if fht_incoherence_maj_bdr_count > 0 else 0.0
        percentage_incoherence_maj_bdr_count = round(percentage_incoherence_maj_bdr_count)

        percentage_incoherence_doublons_ip_bdr_count = (fht_incoherence_doublons_ip_bdr_count / total_records_with_tdj_cdr) * 100 if fht_incoherence_doublons_ip_bdr_count > 0 else 0.0
        percentage_incoherence_doublons_ip_bdr_count = round(percentage_incoherence_doublons_ip_bdr_count)

        percentage_incoherence_nc_swap_fh_bdr_count = (fht_incoherence_nc_swap_fh_bdr_count / total_records_with_tdj_cdr) * 100 if fht_incoherence_nc_swap_fh_bdr_count > 0 else 0.0
        percentage_incoherence_nc_swap_fh_bdr_count = round(percentage_incoherence_nc_swap_fh_bdr_count)

        percentage_incoherence_bde = (fht_incoherence_bde_count / total_records_with_tdj_cdr) * 100 if fht_incoherence_bde_count > 0 else 0.0
        percentage_incoherence_bde = round(percentage_incoherence_bde)

        percentage_incoherence_bdr_ou_bde_verification_trafic_count = (fht_incoherence_bdr_ou_bde_verification_trafic_count / total_records_with_tdj_cdr) * 100 if fht_incoherence_bdr_ou_bde_verification_trafic_count > 0 else 0.0
        percentage_incoherence_bdr_ou_bde_verification_trafic_count = round(percentage_incoherence_bdr_ou_bde_verification_trafic_count)

        percentage_incoherence_bdr_ou_bde_ip_incoherence_count = (fht_incoherence_bdr_ou_bde_ip_incoherence_count / total_records_with_tdj_cdr) * 100 if fht_incoherence_bdr_ou_bde_ip_incoherence_count > 0 else 0.0
        percentage_incoherence_bdr_ou_bde_ip_incoherence_count = round(percentage_incoherence_bdr_ou_bde_ip_incoherence_count)

        percentage_incoherence_swap = (fht_incoherence_swap_count / total_records_with_tdj_cdr) * 100 if fht_incoherence_swap_count > 0 else 0.0
        percentage_incoherence_swap = round(percentage_incoherence_swap)
        
        percentage_incoherence_fhtadeployer = (fht_incoherence_adeployer_count / total_records_with_tdj_cdr) * 100 if fht_incoherence_adeployer_count > 0 else 0.0
        percentage_incoherence_fhtadeployer = round(percentage_incoherence_fhtadeployer)

        result_data = {
            'Pourcentage Incoherence Necessitant MAJ BDR du Parc': percentage_incoherence_maj_bdr_count,
            'Pourcentage Incoherence Doublons Ip BDR du Parc': percentage_incoherence_doublons_ip_bdr_count,
            'Pourcentage Incoherence NC SWAP FH BDR du Parc':percentage_incoherence_nc_swap_fh_bdr_count,
            #'percentage_incoherence_bde': percentage_incoherence_bde,
            #'percentage_incoherence_fhtadeployer': percentage_incoherence_fhtadeployer,
            #'percentage_incoherence_bdr_ou_bde_verification_trafic': percentage_incoherence_bdr_ou_bde_verification_trafic_count,
            #'percentage_incoherence_bdr_ou_bde_ip_incoherence': percentage_incoherence_bdr_ou_bde_ip_incoherence_count,
            #'percentage_incoherence_swap': percentage_incoherence_swap,
            #'total_records': total_records_with_tdj_cdr,
            #'fht_incoherence_maj_bdr_count': fht_incoherence_maj_bdr_count,
            #'fht_incoherence_doublons_ip_bdr_count':fht_incoherence_doublons_ip_bdr_count,
            #'fht_incoherence_nc_swap_fh_bdr_count':fht_incoherence_nc_swap_fh_bdr_count,
            #'fht_incoherence_bde_count': fht_incoherence_bde_count,
            #'fht_incoherence_bdr_ou_bde_verification_trafic_count': fht_incoherence_bdr_ou_bde_verification_trafic_count,
            #'percentage_incoherence_bdr_ou_bde_ip_incoherence_count': percentage_incoherence_bdr_ou_bde_ip_incoherence_count,
            #'fht_incoherence_swap_count':fht_incoherence_swap_count,
            #'fht_incoherence_adeployer_count':fht_incoherence_adeployer_count,
        }

        return JsonResponse(result_data)

class AuditIpFhIncoherenceTdjCdrendetailsBdrouBde(APIView):
    def get(self, request):
        # Count the total number of rows
          # Calculate the date range for the current month
        today = date.today()
        current_month_start = today.replace(day=1)
        current_month_end = today

        # Count the total number of rows with tdj_cdr='TDJ CDR' for the current month
        total_records_with_tdj_cdr = Audit_ip_fh_model.objects.filter(
            tdj_cdr='TDJ CDR',
            Data_Insertion_Date__range=(current_month_start, current_month_end)
        ).count()

        # Count the number of rows for 'Vérification du trafic et mise à jour de BDR ou BDE',
        # 'Adresse IP incohérente entre BDR & BDE' with tdj_cdr='TDJ CDR' for the current month
        fht_incoherence_bdr_ou_bde_verification_trafic_count = Audit_ip_fh_model.objects.filter(
            Q(action_description='Vérification du trafic et mise à jour de BDR ou BDE'),
            tdj_cdr='TDJ CDR',
            Data_Insertion_Date__range=(current_month_start, current_month_end)
        ).count()
        
        fht_incoherence_bdr_ou_bde_ip_incoherence_count = Audit_ip_fh_model.objects.filter(
            Q(action_description='Adresse IP incohérente entre BDR & BDE'),
            tdj_cdr='TDJ CDR',
            Data_Insertion_Date__range=(current_month_start, current_month_end)
        ).count()

        # Calculate the percentage
       
        

        percentage_incoherence_bdr_ou_bde_verification_trafic_count = (fht_incoherence_bdr_ou_bde_verification_trafic_count / total_records_with_tdj_cdr) * 100 if fht_incoherence_bdr_ou_bde_verification_trafic_count > 0 else 0.0
        percentage_incoherence_bdr_ou_bde_verification_trafic_count = round(percentage_incoherence_bdr_ou_bde_verification_trafic_count)

        percentage_incoherence_bdr_ou_bde_ip_incoherence_count = (fht_incoherence_bdr_ou_bde_ip_incoherence_count / total_records_with_tdj_cdr) * 100 if fht_incoherence_bdr_ou_bde_ip_incoherence_count > 0 else 0.0
        percentage_incoherence_bdr_ou_bde_ip_incoherence_count = round(percentage_incoherence_bdr_ou_bde_ip_incoherence_count)

        result_data = {
            'Pourcentage Incoherence BDR ou BDE Necessitant Verification Trafic': percentage_incoherence_bdr_ou_bde_verification_trafic_count,
            'Pourcentage Incoherence BDR ou BDE Ip Incoherente': percentage_incoherence_bdr_ou_bde_ip_incoherence_count,
 
        }

        return JsonResponse(result_data)

class CountPorteurOccurrencesTdjCdrView(APIView):
    def get(self, request):
        # Get the maximum "Data_Insertion_Date" value to determine the latest month
        latest_data_insertion_date = Audit_ip_fh_model.objects.aggregate(Max('Data_Insertion_Date'))['Data_Insertion_Date__max']
        
        if latest_data_insertion_date is not None:
            # Extract the year and month from the latest date
            year = latest_data_insertion_date.year
            month = latest_data_insertion_date.month

            # Create a datetime object for the first day of the latest month
            latest_month_start = datetime(year, month, 1)

            # Filter records where "tdj_cdr" contains "TDJ CDR" and "Data_Insertion_Date" is in the latest month
            porteur_values = Audit_ip_fh_model.objects.filter(
                tdj_cdr="TDJ CDR",
                Data_Insertion_Date__gte=latest_month_start
            ).values_list('porteur', flat=True)

            # Filter out blank values
            non_blank_porteur_values = [value for value in porteur_values if value]

            # Count the occurrences of each distinct non-blank value
            porteur_counts = dict(Counter(non_blank_porteur_values))
            print(porteur_counts)
            return JsonResponse(porteur_counts)
        else:
            # Handle the case where there are no records
            return JsonResponse({"message": "No records found for the latest month"}, status=404)

################ Audit CSG ####################

class CSGUploadView(APIView):
    def process_excel_csg(self, file):
        # Load the Excel file into a DataFrame
        df = pd.read_excel(file)

        # Iterate over each row in the DataFrame
        for _, row in df.iterrows():
            # Create a dictionary to hold the data dynamically
            data = {
                'n_csg': row.get('n_csg', ''),
                'usine_prod': row.get('usine_prod', ''),
                'constructeur_csg': row.get('constructeur_csg', ''),
                'architecture': row.get('architecture', ''),
                'connexion': row.get('connexion', ''),
                'equipment_parent': row.get('equipment_parent', ''),
                'equipement_parent_bdr': row.get('equipement_parent_bdr', ''),
                'delta_equipement_parent': row.get('delta_equipement_parent', ''),
                'port_csg': row.get('port_csg', ''),
                'port_csg_bdr': row.get('port_csg_bdr', ''),
                'delta_port_csg': row.get('delta_port_csg', ''),
                'etat_port_csg': row.get('etat_port_csg', ''),
                'lag_csg': row.get('lag_csg', ''),
                'lag_csg_bdr': row.get('lag_csg_bdr', ''),
                'delta_lag_csg': row.get('delta_lag_csg', ''),
                'loopback0_csg': row.get('loopback0_csg', ''),
                'loopback0_csg_bdr': row.get('loopback0_csg_bdr', ''),
                'delta_loopback0_csg': row.get('delta_loopback0_csg', ''),
                'loopback1_csg': row.get('loopback1_csg', ''),
                'loopback1_csg_bdr': row.get('loopback1_csg_bdr', ''),
                'delta_loopback1_csg': row.get('delta_loopback1_csg', ''),
                'ip_interco_csg': row.get('ip_interco_csg', ''),
                'ip_interco_csg_bdr': row.get('ip_interco_csg_bdr', ''),
                'delta_ip_interco_csg': row.get('delta_ip_interco_csg', ''),
                'vlan_interco_csg': row.get('vlan_interco_csg', ''),
                'vlan_interco_csg_bdr': row.get('vlan_interco_csg_bdr', ''),
                'delta_vlan_trafic_csg': row.get('delta_vlan_trafic_csg', ''),
                'ip_sup_csg': row.get('ip_sup_csg', ''),
                'ip_interco_sup_csg_bdr': row.get('ip_interco_sup_csg_bdr', ''),
                'delta_ip_interco_sup_csg_bdr': row.get('delta_ip_interco_sup_csg_bdr', ''),
                'vlan_sup_csg': row.get('vlan_sup_csg', ''),
                'vlan_interco_sup_csg_bdr': row.get('vlan_interco_sup_csg_bdr', ''),
                'delta_vlan_interco_sup_csg_bdr': row.get('delta_vlan_interco_sup_csg_bdr', ''),
                'port_parent': row.get('port_parent', ''),
                'port_ctr_bdr': row.get('port_ctr_bdr', ''),
                'delta_port_ctr': row.get('delta_port_ctr', ''),
                'lag_parent': row.get('lag_parent', ''),
                'etat_port_parent': row.get('etat_port_parent', ''),
                'lag_ctr_bdr': row.get('lag_ctr_bdr', ''),
                'delta_lag_ctr': row.get('delta_lag_ctr', ''),
                'ip_interco_parent': row.get('ip_interco_parent', ''),
                'ip_interco_ctr_bdr': row.get('ip_interco_ctr_bdr', ''),
                'delta_ip_interco_ctr_bdr': row.get('delta_ip_interco_ctr_bdr', ''),
                'vlan_interco_parent': row.get('vlan_interco_parent', ''),
                'vlan_interco_ctr_bdr': row.get('vlan_interco_ctr_bdr', ''),
                'delta_vlan_interco_ctr_bdr': row.get('delta_vlan_interco_ctr_bdr', ''),
                'ip_sup_parent': row.get('ip_sup_parent', ''),
                'ip_interco_sup_ctr_bdr': row.get('ip_interco_sup_ctr_bdr', ''),
                'delta_ip_interco_sup_ctr_bdr': row.get('delta_ip_interco_sup_ctr_bdr', ''),
                'vlan_sup_parent': row.get('vlan_sup_parent', ''),
                'vlan_sup_ctr_bdr': row.get('vlan_sup_ctr_bdr', ''),
                'delta_vlan_sup_ctr_bdr': row.get('delta_vlan_sup_ctr_bdr', ''),
                'ip_radio': row.get('ip_radio', ''),
                'ip_radio_bdr': row.get('ip_radio_bdr', ''),
                'delta_ip_radio_bdr': row.get('delta_ip_radio_bdr', ''),
                'port_radio': row.get('port_radio', ''),
                'port_radio_bdr': row.get('port_radio_bdr', ''),
                'delta_port_radio_bdr': row.get('delta_port_radio_bdr', ''),
                'lag_radio': row.get('lag_radio', ''),
                'lag_radio_bdr': row.get('lag_radio_bdr', ''),
                'delta_lag_radio_bdr': row.get('delta_lag_radio_bdr', ''),
                'vlan_radio': row.get('vlan_radio', ''),
                'vlan_radio_bdr': row.get('vlan_radio_bdr', ''),
                'delta_vlan_radio_bdr': row.get('delta_vlan_radio_bdr', ''),
                'etat_bdr_bde': row.get('etat_bdr_bde', ''),
                'lien_ip_interco_ctr1': row.get('lien_ip_interco_ctr1', ''),
                'routage_lien_ip_ctr1': row.get('routage_lien_ip_ctr1', ''),
                'lien_ip_sup_ctr1': row.get('lien_ip_sup_ctr1', ''),
                'routage_lien_ip_sup_ctr1': row.get('routage_lien_ip_sup_ctr1', ''),
                'lien_agg_ctr1': row.get('lien_agg_ctr1', ''),
                'routage_lien_agg_ctr1': row.get('routage_lien_agg_ctr1', ''),
                'lien_zt_ctr1': row.get('lien_zt_ctr1', ''),
                'routage_lien_zt_ctr1': row.get('routage_lien_zt_ctr1', ''),
                'lien_ip_interco_ctr2': row.get('lien_ip_interco_ctr2', ''),
                'routage_lien_ip_interco_ctr2': row.get('routage_lien_ip_interco_ctr2', ''),
                'lien_ip_sup_ctr2': row.get('lien_ip_sup_ctr2', ''),
                'routage_lien_ip_interco_sup_ctr2': row.get('routage_lien_ip_interco_sup_ctr2', ''),
                'lien_zt_ctr2': row.get('lien_zt_ctr2', ''),
                'routage_lien_zt_ctr2': row.get('routage_lien_zt_ctr2', ''),
                'lien_agg_ctr2': row.get('lien_agg_ctr2', ''),
                'routage_lien_agg_ctr2': row.get('routage_lien_agg_ctr2', ''),
                'lien_lo_ctr1': row.get('lien_lo_ctr1', ''),
                'lien_lo_ctr2': row.get('lien_lo_ctr2', ''),
                'lag': row.get('lag', ''),
                'lag_global': row.get('lag_global', ''),
                'vlan': row.get('vlan', ''),
                'vlan_global': row.get('vlan_global', ''),
                'adresse_ip_interco': row.get('adresse_ip_interco', ''),
                'adresse_ip_sup': row.get('adresse_ip_sup', ''),
                'adresse_ip_loopback0_1': row.get('adresse_ip_loopback0_1', ''),
                'adresse_ip_radio': row.get('adresse_ip_radio', ''),
                'modelisation_des_liens': row.get('modelisation_des_liens', ''),
                'port_csg_ctr': row.get('port_csg_ctr', ''),
                'port_global': row.get('port_global', ''),
                'routage_des_liens': row.get('routage_des_liens', ''),
                'etat_de_coherence_imes': row.get('etat_de_coherence_imes', ''),
                'etat_de_coherence_radio': row.get('etat_de_coherence_radio', ''),
                'etat_csg_globale': row.get('etat_csg_globale', ''),
                'insertion_date': row.get('insertion_date', ''),
            }

            # Save data into the AuditCSG model
            AuditCSG.objects.create(**data)

        return {"message": "Data processed and saved successfully!"}

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        if not file:
            return Response({"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Process the file
            result = self.process_excel_csg(file)
            return Response(result, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class AuditCSGRadio(APIView):
    def getLatestDate(self, architecture):
        latest_record = (
            AuditCSG.objects.filter(architecture=architecture).order_by('-insertion_date').values('insertion_date').first()
        )
        # If no records exist, return an empty response or handle as needed
        if not latest_record:
            return None
        
        latest_month = latest_record['insertion_date']
        return latest_month

    def tauxdecoherence(self, architecture):
        # Determine the latest month
        latest_month = self.getLatestDate(architecture)

        # Count the total number of rows with the latest month
        total_records = AuditCSG.objects.filter(
            insertion_date=latest_month
        ).count()

        # Count the number of rows with 'CSG-FH-CTR' in the Architecture column for the latest month
        csg_count = AuditCSG.objects.filter(
            architecture= architecture,
            insertion_date=latest_month
        ).values('n_csg').distinct().count()

        # Calculate the percentage
        percentage = (csg_count / 10185) * 100 if total_records > 0 else 0.0
        percentage = round(percentage)
        result_data = {
            'percentage': percentage,
            'total_records': total_records,
            'csg_count': csg_count,
        }

        return result_data

    def repartitionCsg(self, architecture):
        # Get all records with value 'Site avec au moins une incohérence' in ETAT_CSG
        latest_date =  self.getLatestDate(architecture)
        records = AuditCSG.objects.filter(architecture=architecture, etat_de_coherence_radio='Site avec au moins une incohérence', insertion_date=latest_date)
        
        # Columns to check for NOK values
        columns_to_check = [
            'adresse_ip_radio','delta_lag_radio_bdr','delta_vlan_radio_bdr','delta_port_radio_bdr'
        ]

        # Initialize counts
        one_nok_count = 0
        two_nok_count = 0
        three_or_more_nok_count = 0

        # Iterate over each record
        for record in records:
            nok_count = sum(1 for column in columns_to_check if getattr(record, column) == 'NOK')
            
            if nok_count == 1:
                one_nok_count += 1
            elif nok_count == 2:
                two_nok_count += 1
            elif nok_count > 2:
                three_or_more_nok_count += 1

        # Return the counts as JSON response
        data = {
            'one_nok_count': one_nok_count,
            'two_nok_count': two_nok_count,
            'three_or_more_nok_count': three_or_more_nok_count,
        }
        return data

    def calculate_nok_percentage(self, architecture):
        # Define display names for columns
        display_names = {
            'adresse_ip_radio': 'Address IP Radio',
            'delta_lag_radio_bdr': 'LAG',
            'delta_vlan_radio_bdr': 'VLAN',
            'delta_port_radio_bdr': 'Port'
        }

        # Columns to check for straightforward NOK checks
        columns = list(display_names.keys())

        # Initialize NOK counters
        nok_counts = {column: 0 for column in columns}
        total_nok_count = 0

        # Get the latest date for the provided architecture
        latest_date = self.getLatestDate(architecture)

        # Fetch records for the given architecture at the latest date
        records = AuditCSG.objects.filter(architecture=architecture, insertion_date=latest_date)

        # Count NOKs for each column
        for record in records:
            for column in columns:
                if getattr(record, column, 'OK') == 'NOK':
                    nok_counts[column] += 1
                    total_nok_count += 1

        # If no NOKs found, return zero for all fields with display names
        if total_nok_count == 0:
            return {display_names[column]: 0 for column in columns}

        # Calculate and return counts with display names
        return {display_names[column]: count for column, count in nok_counts.items()}

    def evolutionDeCoherance(self, architecture):
        # 1. Fetch the latest insertion_date for this architecture
        latest_record = (
            AuditCSG.objects
            .filter(architecture=architecture)
            .order_by('-insertion_date')
            .values('insertion_date')
            .first()
        )

        # If no records exist, return an empty list
        if not latest_record:
            return []

        # latest_record is something like: {'insertion_date': datetime(2023, 8, 15, ...)}
        latest_insertion_date = latest_record['insertion_date']
        
        # 2. Determine the start/end of the latest month
        latest_month_start = latest_insertion_date.replace(day=1)
        latest_month_end = latest_month_start + relativedelta(months=1, days=-1)
        
        # 3. Determine the start/end of the previous month
        #    (by going one day before `latest_month_start` to find the end of the previous month)
        previous_month_end = latest_month_start - timedelta(days=1)
        previous_month_start = previous_month_end.replace(day=1)

        # 4. Query for the LATEST month
        grouped_data_latest_month = (
            AuditCSG.objects.filter(
                insertion_date__range=(latest_month_start, latest_month_end),
                architecture=architecture
            )
            .values('insertion_date')
            .annotate(
                total_records=Count('n_csg'),
                sites_conformes=Count(
                    Case(When(etat_de_coherence_radio='Site conforme', then=Value(1)))
                )
            )
        )

        # 5. Query for the PREVIOUS month
        grouped_data_previous_month = (
            AuditCSG.objects.filter(
                insertion_date__range=(previous_month_start, previous_month_end),
                architecture=architecture
            )
            .values('insertion_date')
            .annotate(
                total_records=Count('n_csg'),
                sites_conformes=Count(
                    Case(When(etat_de_coherence_radio='Site conforme', then=Value(1)))
                )
            )
        )

        # 6. Calculate percentages and differences
        result_data = []

        # We zip these two QuerySets so we can compare them month-by-month
        for item_latest, item_previous in zip(grouped_data_latest_month, grouped_data_previous_month):
            total_records_latest = item_latest['total_records']
            sites_conformes_latest = item_latest['sites_conformes']
            percentage_latest = (
                (sites_conformes_latest / total_records_latest) * 100
                if total_records_latest > 0
                else 0.0
            )
            percentage_latest = round(percentage_latest)

            total_records_previous = item_previous['total_records']
            sites_conformes_previous = item_previous['sites_conformes']
            percentage_previous = (
                (sites_conformes_previous / total_records_previous) * 100
                if total_records_previous > 0
                else 0.0
            )

            difference = percentage_latest - percentage_previous
            difference = round(difference)

            result_data.append({
                # Convert insertion_date to string for serialization
                'date': item_latest['insertion_date'].strftime("%Y-%m-%d"),
                'difference': abs(difference),
                'positive': difference >= 0,
                'value': f"{round(percentage_latest)}%", 
                'category': self.calculate_category(percentage_latest),
            })
        

        return result_data

    def calculate_category(self, percentage):
        if percentage < 60:
            return 'low'
        elif 60 <= percentage <= 90:
            return 'medium'
        else:
            return 'high'

    def evolution(self,architecture):
        filtered_data = AuditCSG.objects.filter(architecture=architecture)
        
        grouped_data = filtered_data.annotate(month=TruncMonth('insertion_date')).values('month') \
            .annotate(
                sites_conformes=Count(Case(When(etat_de_coherence_radio='Site conforme', then=1))),
                sites_avec_incoherence=Count(Case(When(etat_de_coherence_radio='Site avec au moins une incohérence', then=1)))
            ).order_by('-month')[:11]

        return grouped_data

    def get(self, request, architecture, format=None):
            # Extract the architecture from the query parameters (if provided)
            try:
                # Fetch data from each method
                latest_date = self.getLatestDate(architecture)

                if not latest_date:
                    # If there is no record, handle however you like
                    # e.g., return an empty object or custom message
                    return Response(
                        {
                            "latest_date": None,
                            "message": f"No records found for architecture={architecture}"
                        },
                        status=status.HTTP_200_OK
                    )


                taux_de_coherence = self.tauxdecoherence(architecture)
                repartition_data = self.repartitionCsg(architecture)
                nok_percentage = self.calculate_nok_percentage(architecture)
                evolution_de_coherance = self.evolutionDeCoherance(architecture)
                evolution_data = self.evolution(architecture)

                # Prepare the final response
                response_data = {
                    'latest_date': latest_date,
                    'taux_de_coherence': taux_de_coherence,
                    'repartition_data': repartition_data,
                    'nok_percentage': nok_percentage,
                    'evolution_de_coherance': evolution_de_coherance,
                    'evolution': list(evolution_data),  # Convert queryset to list for serialization
                }

                return Response(response_data, status=status.HTTP_200_OK)

            except Exception as e:
                return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class AuditCSGImes(APIView):
    def getLatestDate(self, architecture):
        latest_record = (
            AuditCSG.objects.filter(architecture=architecture).order_by('-insertion_date').values('insertion_date').first()
        )
        # If no records exist, return an empty response or handle as needed
        if not latest_record:
            # Return None (or a Python date object, or an empty string)
            return None
        
        latest_month = latest_record['insertion_date']

        return latest_month

    def tauxdecoherence(self, architecture):
        # Determine the latest month
        latest_month = self.getLatestDate(architecture)

        # Count the total number of rows with the latest month
        total_records = AuditCSG.objects.filter(
            insertion_date=latest_month
        ).count()

        # Count the number of rows with 'CSG-FH-CTR' in the Architecture column for the latest month
        csg_count = AuditCSG.objects.filter(
            architecture= architecture,
            insertion_date=latest_month
        ).values('n_csg').distinct().count()

        # Calculate the percentage
        percentage = (csg_count / 10185) * 100 if total_records > 0 else 0.0
        percentage = round(percentage)
        result_data = {
            'percentage': percentage,
            'total_records': total_records,
            'csg_count': csg_count,
        }

        return result_data

    def repartitionCsg(self, architecture):
        # Get all records with value 'Site avec au moins une incohérence' in ETAT_CSG
        latest_date = self.getLatestDate(architecture)
        # records = AuditCSG.objects.filter(architecture=architecture, etat_de_coherence_imes='Site avec au moin une incohérence', insertion_date=latest_date)
        records = (
            AuditCSG.objects
                .filter(architecture=architecture, insertion_date=latest_date)  # keep your existing filters
                .exclude(etat_de_coherence_imes='Site Conforme')                # drop anything “Site Conforme”
        )
        # Columns to check for NOK values
        columns_to_check = [
            'adresse_ip_interco', 'adresse_ip_sup', 'adresse_ip_loopback0_1', 
            'modelisation_des_liens', 'routage_des_liens','lag', 'vlan', 'port_csg_ctr'
        ]

        # Initialize counts
        one_nok_count = 0
        two_nok_count = 0
        three_or_more_nok_count = 0

        # Iterate over each record
        for record in records:
            nok_count = sum(1 for column in columns_to_check if getattr(record, column) == 'NOK')
            
            if nok_count == 1:
                one_nok_count += 1
            elif nok_count == 2:
                two_nok_count += 1
            elif nok_count >= 3:
                three_or_more_nok_count += 1

        # Return the counts as JSON response
        data = {
            'one_nok_count': one_nok_count,
            'two_nok_count': two_nok_count,
            'three_or_more_nok_count': three_or_more_nok_count,
        }

        return data

    def calculate_nok_percentage(self, architecture):
        latest_month = self.getLatestDate(architecture)

        # Mapping column names to user-friendly labels
        column_labels = {
            'adresse_ip_interco': 'Addresse IP Interco',
            'adresse_ip_sup': 'Addresse IP SUP',
            'adresse_ip_loopback0_1': 'Loopback0/1 Addresse IP',
            'modelisation_des_liens': 'Modelisation des liens',
            'routage_des_liens': 'Routage des liens',
            'lag': 'LAG',
            'vlan': 'VLAN',
            'port_csg_ctr': 'Port'
        }

        # Initialize counts
        nok_counts = {column: 0 for column in column_labels}
        total_nok_count = 0

        # Fetch records for the given architecture
        records = AuditCSG.objects.filter(architecture=architecture, insertion_date=latest_month)

        # Iterate through each record to calculate NOK counts
        for record in records:
            # Check and count NOKs for standard columns
            for column in column_labels:
                if getattr(record, column, 'NOK') == 'NOK':
                    nok_counts[column] += 1
                    total_nok_count += 1

        # If there are no NOKs, return zero for all fields
        if total_nok_count == 0:
            return {label: 0 for label in column_labels.values()}

        # Calculate percentages for each column
        nok_percentages = {}
        for column, count in nok_counts.items():
            label = column_labels[column]
            nok_percentages[label] = round(count)

        return nok_percentages

    def evolutionDeCoherance(self, architecture):
        # 1. Fetch the latest insertion_date for this architecture
        latest_record = (
            AuditCSG.objects
            .filter(architecture=architecture)
            .order_by('-insertion_date')
            .values('insertion_date')
            .first()
        )

        if not latest_record:
            return []

        latest_insertion_date = latest_record['insertion_date']
        
        # 2. Determine the start/end of the latest month
        latest_month_start = latest_insertion_date.replace(day=1)
        latest_month_end = latest_month_start + relativedelta(months=1, days=-1)
        
        # 3. Determine the start/end of the previous month
        previous_month_end = latest_month_start - timedelta(days=1)
        previous_month_start = previous_month_end.replace(day=1)

        # 4. Query for the LATEST month
        grouped_data_latest_month = (
            AuditCSG.objects.filter(
                insertion_date__range=(latest_month_start, latest_month_end),
                architecture=architecture
            )
            .values('insertion_date')
            .annotate(
                total_records=Count('n_csg'),
                sites_conformes=Count(
                    Case(When(etat_de_coherence_imes='Site conforme', then=Value(1)))
                )
            )
        )

        # 5. Query for the PREVIOUS month
        grouped_data_previous_month = (
            AuditCSG.objects.filter(
                insertion_date__range=(previous_month_start, previous_month_end),
                architecture=architecture
            )
            .values('insertion_date')
            .annotate(
                total_records=Count('n_csg'),
                sites_conformes=Count(
                    Case(When(etat_de_coherence_imes='Site conforme', then=Value(1)))
                )
            )
        )

        # 6. Calculate percentages and differences
        result_data = []

        # We zip these two QuerySets so we can compare them month-by-month
        for item_latest, item_previous in zip(grouped_data_latest_month, grouped_data_previous_month):
            total_records_latest = item_latest['total_records']
            sites_conformes_latest = item_latest['sites_conformes']
            percentage_latest = (
                (sites_conformes_latest / total_records_latest) * 100
                if total_records_latest > 0
                else 0.0
            )
            percentage_latest = round(percentage_latest)

            total_records_previous = item_previous['total_records']
            sites_conformes_previous = item_previous['sites_conformes']
            percentage_previous = (
                (sites_conformes_previous / total_records_previous) * 100
                if total_records_previous > 0
                else 0.0
            )

            difference = percentage_latest - percentage_previous
            difference = round(difference)

            result_data.append({
                # Convert insertion_date to string for serialization
                'date': item_latest['insertion_date'].strftime("%Y-%m-%d"),
                'difference': abs(difference),
                'positive': difference >= 0,
                'value': f"{round(percentage_latest)}%", 
                'category': self.calculate_category(percentage_latest),
            })

        return result_data

    def calculate_category(self, percentage):
        if percentage < 60:
            return 'low'
        elif 60 <= percentage <= 90:
            return 'medium'
        else:
            return 'high'

    def evolution(self,architecture):
        filtered_data = AuditCSG.objects.filter(architecture=architecture)
        
        grouped_data = filtered_data.annotate(month=TruncMonth('insertion_date')).values('month') \
            .annotate(
                sites_conformes=Count(Case(When(etat_de_coherence_imes='Site conforme', then=1))),
                sites_avec_incoherence = Count(
                    Case(
                        When(~Q(etat_de_coherence_imes='Site conforme'), then=1),
                    )
                )            
            ).order_by('-month')[:11]

        return grouped_data

    def get(self, request, architecture, format=None):
        # Extract the architecture from the query parameters (if provided)
        try:
            # Fetch data from each method
            latest_date = self.getLatestDate(architecture)

            if not latest_date:
                # If there is no record, handle however you like
                return Response(
                    {
                        "latest_date": None,
                        "message": f"No records found for architecture={architecture}"
                    },
                    status=status.HTTP_200_OK
                )

            taux_de_coherence = self.tauxdecoherence(architecture)
            repartition_data = self.repartitionCsg(architecture)
            nok_percentage = self.calculate_nok_percentage(architecture)
            evolution_de_coherance = self.evolutionDeCoherance(architecture)
            evolution_data = self.evolution(architecture)

            # Prepare the final response
            response_data = {
                'latest_date': latest_date,
                'taux_de_coherence': taux_de_coherence,
                'repartition_data': repartition_data,
                'nok_percentage': nok_percentage,
                'evolution_de_coherance': evolution_de_coherance,
                'evolution': list(evolution_data),  # Convert queryset to list for serialization
            }

            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class AuditCSGRadioGlobal(APIView):
    def getLatestDate(self):
        latest_record = (
            AuditCSG.objects.order_by('-insertion_date').values('insertion_date').first()
        )
        # If no records exist, return an empty response or handle as needed
        if not latest_record:
            # Return None (or a Python date object, or an empty string)
            return None
        
        latest_month = latest_record['insertion_date']

        return latest_month

    def tauxdecoherence(self):
        # Determine the latest month
        latest_month = AuditCSG.objects.aggregate(
            latest_month=Max(ExtractMonth('insertion_date'))
        )['latest_month']

        # Count the total number of rows with the latest month
        total_records = AuditCSG.objects.filter(
            insertion_date__month=latest_month
        ).count()

        # Count the number of rows with 'CSG-FH-CTR' in the Architecture column for the latest month
        csg_count = AuditCSG.objects.filter(
            insertion_date__month=latest_month
        ).values('n_csg').distinct().count()

        # Calculate the percentage
        percentage =  100 if total_records > 0 else 0.0
        percentage = round(percentage)
        result_data = {
            'percentage': percentage,
            'total_records': total_records,
            'csg_count': csg_count,
        }

        return result_data

    def repartitionCsg(self):
        # Get all records with value 'Site avec au moins une incohérence' in ETAT_CSG
        latest_date = AuditCSG.objects.filter(etat_de_coherence_radio='Site avec au moins une incohérence').aggregate(Max('insertion_date'))['insertion_date__max']
        records = AuditCSG.objects.filter(etat_de_coherence_radio='Site avec au moins une incohérence', insertion_date=latest_date)
        
        # Columns to check for NOK values
        columns_to_check = [
            'adresse_ip_radio','delta_lag_radio_bdr','delta_vlan_radio_bdr','delta_port_radio_bdr'
        ]

        # Initialize counts
        one_nok_count = 0
        two_nok_count = 0
        three_or_more_nok_count = 0

        # Iterate over each record
        for record in records:
            nok_count = sum(1 for column in columns_to_check if getattr(record, column) == 'NOK')
            
            if nok_count == 1:
                one_nok_count += 1
            elif nok_count == 2:
                two_nok_count += 1
            elif nok_count >= 3:
                three_or_more_nok_count += 1

        # Return the counts as JSON response
        data = {
            'one_nok_count': one_nok_count,
            'two_nok_count': two_nok_count,
            'three_or_more_nok_count': three_or_more_nok_count,
        }
        return data

    def calculate_nok_percentage(self):
        # Dictionary to store the global percentage data
        global_usine_prod_percentages = {}

        # Get the latest date across all records
        latest_date = AuditCSG.objects.aggregate(latest_date=Max('insertion_date'))['latest_date']
        
        if not latest_date:
            return global_usine_prod_percentages  # Return empty if there's no data

        # Filter records for the latest date
        filtered_data = AuditCSG.objects.filter(insertion_date=latest_date)

        # Aggregate the count of each unique value in the Usine_Prod column
        usine_prod_counts = filtered_data.values('usine_prod').annotate(count=Count('usine_prod'))

        # Calculate the total number of records
        total_count = sum([item['count'] for item in usine_prod_counts])

        # Calculate the percentage for each unique value
        if total_count > 0:
            global_usine_prod_percentages = {
                item['usine_prod']: round((item['count'] / total_count) * 100)
                for item in usine_prod_counts
            }
        return global_usine_prod_percentages

    def calculate_nok_ip_routage_percentage(self):
        # Global columns to check NOKs
        global_columns = [
            'adresse_ip_radio', 'lag_radio', 'vlan_radio', 'port_radio'
        ]

        # Specific columns
        imes = [
            'lag', 'vlan', 'port_csg_ctr'
        ]
        globale = [
            'lag_global', 'vlan_global', 'port_global'
        ]
        
        # Specific columns for percentage calculation
        percentage_columns = [
            'vlan_radio', 'adresse_ip_radio'
        ]
        
        total_nok_count = 0
        specific_nok_count = 0
        architecture_nok_counts = {}

        # Get all distinct architectures
        architectures = AuditCSG.objects.values_list('architecture', flat=True).distinct()

        for architecture in architectures:
            # Initialize dictionary for storing NOK counts for this architecture
            architecture_nok_counts[architecture] = {column: 0 for column in global_columns}

            # Get the latest date for the current architecture
            latest_date = AuditCSG.objects.filter(architecture=architecture).aggregate(latest_date=Max('insertion_date'))['latest_date']
            
            if not latest_date:
                continue

            # Filter records for the current architecture and latest date
            filtered_data = AuditCSG.objects.filter(architecture=architecture, insertion_date=latest_date)

            # Apply cascading NOK logic for lag_radio, vlan_radio, and port_radio
            for global_col, specific_col in zip(globale, imes):
                nok_records = filtered_data.filter(
                    **{
                        f"{global_col}": 'NOK',   # Check if global column is NOK
                        f"{specific_col}": 'OK'  # Check if specific column is OK
                    }
                )
                for record in nok_records:
                    # Mark the higher-level column as NOK
                    higher_level_col = global_col.replace('_global', '_radio')
                    architecture_nok_counts[architecture][higher_level_col] += 1
                    total_nok_count += 1

            # Calculate NOK counts for global columns
            for column in global_columns:
                nok_count = filtered_data.filter(**{column: 'NOK'}).count()
                architecture_nok_counts[architecture][column] += nok_count
                total_nok_count += nok_count

            # Calculate total 'NOK' counts for the specific percentage columns for this architecture
            for column in percentage_columns:
                specific_nok_count += architecture_nok_counts[architecture][column]

        # Calculate the percentage of NOKs in the specific columns relative to the global total
        if total_nok_count == 0:
            ip_nok_percentage = 0
            routage_nok_percentage = 0
        else:
            ip_nok_percentage = round((specific_nok_count / total_nok_count) * 100)
            routage_nok_count = total_nok_count - specific_nok_count
            routage_nok_percentage = round((routage_nok_count / total_nok_count) * 100)

        return {
            'Incoherence Routage': ip_nok_percentage,
            'Incoherence IP': routage_nok_percentage,
        }

    def evolutionDeCoherance(self, evolution_data):
        result_data = []

        # Ensure the data is sorted by month
        sorted_data = sorted(evolution_data, key=lambda x: x['month'])

        # Take only the last two months
        if len(sorted_data) >= 2:
            last_two_months = sorted_data[-2:]
        else:
            # Not enough data to calculate the evolution
            return result_data

        # Calculate the percentage difference
        previous_month = last_two_months[0]
        current_month = last_two_months[1]

        # Calculate total sites for current and previous months
        total_current = current_month['sites_conformes'] + current_month['sites_avec_incoherence']
        total_previous = previous_month['sites_conformes'] + previous_month['sites_avec_incoherence']

        # Calculate percentage of conformes for current and previous months
        percentage_current = (current_month['sites_conformes'] / total_current) * 100 if total_current > 0 else 0
        percentage_previous = (previous_month['sites_conformes'] / total_previous) * 100 if total_previous > 0 else 0

        # Calculate the percentage difference
        difference = round(percentage_current - percentage_previous)

        # Append the result for the current month
        result_data.append({
            'date': current_month['month'],
            'difference': abs(difference),
            'positive': difference >= 0,
            'value': f"{round(percentage_current)}%",  # Format the percentage
            'category': self.calculate_category(percentage_current),  # Assuming this method exists
        })

        return result_data
    
    def calculate_category(self, percentage):
        if percentage < 60:
            return 'low'
        elif 60 <= percentage <= 90:
            return 'medium'
        else:
            return 'high'

    def evolution(self):
        filtered_data = AuditCSG.objects.all()
        
        grouped_data = filtered_data.annotate(month=TruncMonth('insertion_date')).values('month') \
            .annotate(
                sites_conformes=Count(Case(When(etat_de_coherence_radio='Site conforme', then=1))),
                sites_avec_incoherence=Count(Case(When(etat_de_coherence_radio='Site avec au moins une incohérence', then=1)))
            ).order_by('-month')[:11]

        return grouped_data

    def get(self, request, format=None):
            # Extract the architecture from the query parameters (if provided)
            try:
                # Fetch data from each method
                latest_date = self.getLatestDate()

                if not latest_date:
                    # If there is no record, handle however you like
                    # e.g., return an empty object or custom message
                    return Response(
                        {
                            "latest_date": None,
                            "message": f"No records found"
                        },
                        status=status.HTTP_200_OK
                    )


                taux_de_coherence = self.tauxdecoherence()
                nok_percentage = self.calculate_nok_percentage()
                nok_ip_routage_percentage = self.calculate_nok_ip_routage_percentage() #keep
                evolution_data = self.evolution() #keep
                evolution_de_coherance = self.evolutionDeCoherance(evolution_data) #keep

                # Prepare the final response
                response_data = {
                    'latest_date': latest_date,
                    'taux_de_coherence': taux_de_coherence,
                    'nok_percentage': nok_percentage,
                    'nok_ip_routage_percentage': nok_ip_routage_percentage,
                    'evolution_de_coherance': evolution_de_coherance,
                    'evolution': list(evolution_data),  # Convert queryset to list for serialization
                }

                return Response(response_data, status=status.HTTP_200_OK)

            except Exception as e:
                return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class AuditCSGImesGlobal(APIView):
    def getLatestDate(self):
        latest_record = (
            AuditCSG.objects.order_by('-insertion_date').values('insertion_date').first()
        )
        # If no records exist, return an empty response or handle as needed
        if not latest_record:
            # Return None (or a Python date object, or an empty string)
            return None
        
        latest_month = latest_record['insertion_date']

        return latest_month

    def tauxdecoherence(self):
        # Determine the latest month
        latest_month = AuditCSG.objects.aggregate(
            latest_month=Max(ExtractMonth('insertion_date'))
        )['latest_month']

        # Count the total number of rows with the latest month
        total_records = AuditCSG.objects.filter(
            insertion_date__month=latest_month
        ).count()

        # Count the number of rows with 'CSG-FH-CTR' in the Architecture column for the latest month
        csg_count = AuditCSG.objects.filter(
            insertion_date__month=latest_month
        ).values('n_csg').distinct().count()

        # Calculate the percentage
        percentage =  100 if total_records > 0 else 0.0
        percentage = round(percentage)
        result_data = {
            'percentage': percentage,
            'total_records': total_records,
            'csg_count': csg_count,
        }

        return result_data

    def repartitionCsg(self):
        # Get all records with value 'Site avec au moins une incohérence' in ETAT_CSG
        latest_date = AuditCSG.objects.aggregate(Max('insertion_date'))['insertion_date__max']
        records = AuditCSG.objects.filter(etat_de_coherence_imes='Site avec au moin une incohérence', insertion_date=latest_date)
        
        # Columns to check for NOK values
        columns_to_check = [
            'adresse_ip_interco', 'adresse_ip_sup', 'adresse_ip_loopback0_1', 
            'modelisation_des_liens', 'routage_des_liens','lag', 'vlan', 'port_csg_ctr'
        ]

        # Initialize counts
        one_nok_count = 0
        two_nok_count = 0
        three_or_more_nok_count = 0

        # Iterate over each record
        for record in records:
            nok_count = sum(1 for column in columns_to_check if getattr(record, column) == 'NOK')
            
            if nok_count == 1:
                one_nok_count += 1
            elif nok_count == 2:
                two_nok_count += 1
            elif nok_count >= 3:
                three_or_more_nok_count += 1

        # Return the counts as JSON response
        data = {
            'one_nok_count': one_nok_count,
            'two_nok_count': two_nok_count,
            'three_or_more_nok_count': three_or_more_nok_count,
        }

        return data

    def calculate_nok_percentage(self):
        # Dictionary to store the global percentage data
        global_usine_prod_percentages = {}

        # Get the latest date across all records
        latest_date = AuditCSG.objects.aggregate(latest_date=Max('insertion_date'))['latest_date']
        
        if not latest_date:
            return global_usine_prod_percentages  # Return empty if there's no data

        # Filter records for the latest date
        filtered_data = AuditCSG.objects.filter(insertion_date=latest_date)

        # Aggregate the count of each unique value in the Usine_Prod column
        usine_prod_counts = filtered_data.values('usine_prod').annotate(count=Count('usine_prod'))

        # Calculate the total number of records
        total_count = sum([item['count'] for item in usine_prod_counts])

        # Calculate the percentage for each unique value
        if total_count > 0:
            global_usine_prod_percentages = {
                item['usine_prod']: round((item['count'] / total_count) * 100)
                for item in usine_prod_counts
            }
        return global_usine_prod_percentages

    def evolutionDeCoherance(self, evolution_data):
        result_data = []

        # Ensure the data is sorted by month
        sorted_data = sorted(evolution_data, key=lambda x: x['month'])

        # Take only the last two months
        if len(sorted_data) >= 2:
            last_two_months = sorted_data[-2:]
        else:
            # Not enough data to calculate the evolution
            return result_data

        # Calculate the percentage difference
        previous_month = last_two_months[0]
        current_month = last_two_months[1]

        # Calculate total sites for current and previous months
        total_current = current_month['sites_conformes'] + current_month['sites_avec_incoherence']
        total_previous = previous_month['sites_conformes'] + previous_month['sites_avec_incoherence']

        # Calculate percentage of conformes for current and previous months
        percentage_current = (current_month['sites_conformes'] / total_current) * 100 if total_current > 0 else 0
        percentage_previous = (previous_month['sites_conformes'] / total_previous) * 100 if total_previous > 0 else 0

        # Calculate the percentage difference
        difference = round(percentage_current - percentage_previous)

        # Append the result for the current month
        result_data.append({
            'date': current_month['month'],
            'difference': abs(difference),
            'positive': difference >= 0,
            'value': f"{round(percentage_current)}%",  # Format the percentage
            'category': self.calculate_category(percentage_current),  # Assuming this method exists
        })

        return result_data

    def calculate_nok_ip_routage_percentage(self):
        # Global columns to check NOKs
        global_columns = [
            'adresse_ip_interco', 'adresse_ip_sup', 'adresse_ip_loopback0_1',
            'modelisation_des_liens', 'routage_des_liens', 'lag', 'vlan', 'port_csg_ctr'
        ]

        # Specific columns for percentage calculation
        percentage_columns = [
            'vlan', 'adresse_ip_interco',
            'adresse_ip_sup', 'adresse_ip_loopback0_1'
        ]
        
        total_nok_count = 0
        specific_nok_count = 0
        architecture_nok_counts = {}

        # Get all distinct architectures
        architectures = AuditCSG.objects.values_list('architecture', flat=True).distinct()

        for architecture in architectures:
            # Initialize dictionary for storing NOK counts for this architecture
            architecture_nok_counts[architecture] = {column: 0 for column in global_columns}

            # Get the latest date for the current architecture
            latest_date = AuditCSG.objects.filter(architecture=architecture).aggregate(latest_date=Max('insertion_date'))['latest_date']
            
            if not latest_date:
                continue

            # Filter records for the current architecture and latest date
            filtered_data = AuditCSG.objects.filter(architecture=architecture, insertion_date=latest_date)

            for column in global_columns:
                nok_count = filtered_data.filter(**{column: 'NOK'}).count()
                architecture_nok_counts[architecture][column] = nok_count
                total_nok_count += nok_count

            # Calculate total 'NOK' counts for the specific percentage columns for this architecture
            for column in percentage_columns:
                specific_nok_count += architecture_nok_counts[architecture][column]

        # Calculate the percentage of NOKs in the specific columns relative to the global total
        if total_nok_count == 0:
            ip_nok_percentage = 0
            routage_nok_percentage = 0
        else:
            ip_nok_percentage = round((specific_nok_count / total_nok_count) * 100)
            routage_nok_count = total_nok_count - specific_nok_count
            routage_nok_percentage = round((routage_nok_count / total_nok_count) * 100)

        return {
            'Incoherence Routage': ip_nok_percentage,
            'Incoherence IP': routage_nok_percentage,
        }

    def calculate_category(self, percentage):
        if percentage < 60:
            return 'low'
        elif 60 <= percentage <= 90:
            return 'medium'
        else:
            return 'high'

    def evolution(self):
        filtered_data = AuditCSG.objects.all()
        
        grouped_data = filtered_data.annotate(month=TruncMonth('insertion_date')).values('month') \
        .annotate(
            sites_conformes=Count(Case(When(etat_de_coherence_imes='Site conforme', then=1))),
            sites_avec_incoherence=Count(Case( When(~Q(etat_de_coherence_imes='Site conforme'), then=1)))            
        ).order_by('-month')[:11]

        return grouped_data

    def get(self, request, format=None):
        # Extract the architecture from the query parameters (if provided)
        try:
            # Fetch data from each method
            latest_date = self.getLatestDate()

            if not latest_date:
                # If there is no record, handle however you like
                # e.g., return an empty object or custom message
                return Response(
                    {
                        "latest_date": None,
                        "message": f"No records found"
                    },
                    status=status.HTTP_200_OK
                )


            taux_de_coherence = self.tauxdecoherence()
            repartition_data = self.repartitionCsg()
            nok_percentage = self.calculate_nok_percentage()
            evolution_data = self.evolution()
            evolution_de_coherance = self.evolutionDeCoherance(evolution_data)
            nok_ip_routage_percentage = self.calculate_nok_ip_routage_percentage()

            # Prepare the final response
            response_data = {
                'latest_date': latest_date,
                'taux_de_coherence': taux_de_coherence,
                'repartition_data': repartition_data,
                'nok_percentage': nok_percentage,
                'nok_ip_routage_percentage': nok_ip_routage_percentage,
                'evolution_de_coherance': evolution_de_coherance,
                'evolution': list(evolution_data),  # Convert queryset to list for serialization
            }

            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class ExportAuditCSG(APIView):

    def get(self, request, architecture=None, format=None):
        # 1. Create an in-memory Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Audit CSG"

        # 2. Dynamically fetch all concrete (non-relational) fields from your model
        model_fields = [
            field.name
            for field in AuditCSG._meta.get_fields()
            if getattr(field, "concrete", False) and not getattr(field, "is_relation", False)
        ]

        # 3. Add your field names as the header row
        ws.append(model_fields)

        if(architecture):
            # 4. Query for data
            latest_record = (
                AuditCSG.objects.filter(architecture=architecture).order_by('-insertion_date').values('insertion_date').first()
            )
            data = AuditCSG.objects.filter(architecture=architecture,insertion_date=latest_record['insertion_date'])
        else:
            latest_record = (
                AuditCSG.objects.order_by('-insertion_date').values('insertion_date').first()
            )
            data = AuditCSG.objects.filter(insertion_date=latest_record['insertion_date'])
        # 5. Loop through each record and append all fields,
        #    converting datetime/time objects with tzinfo to naive.
        for record in data:
            row = []
            for field_name in model_fields:
                value = getattr(record, field_name, None)
                # If the value is a datetime or time and has tzinfo, remove it
                if isinstance(value, (datetime, time)) and getattr(value, "tzinfo", None):
                    value = value.replace(tzinfo=None)
                row.append(value)
            ws.append(row)

        # 6. Build the HTTP response with the correct content type
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="audit_records.xlsx"'

        # 7. Save the workbook to the response
        wb.save(response)
        return response


class ATRUploadView(APIView):

    def process_excel_atr(self, file):
        df = pd.read_excel(file)

        # Iterate over each row in the DataFrame
        for _, row in df.iterrows():
            data = {
                'key': row.get('Key', ''),
                'atr': row.get('ATR', ''),
                'connexion': row.get('connexion', ''),
                'equipment_parent': row.get('equipment_parent', ''),
                'equipement_parent_bdr': row.get('equipement_parent_bdr', ''),
                'delta_equipement_parent': row.get('delta_equipement_parent', ''),
                'port_atr': row.get('port_atr', ''),
                'port_atr_bdr': row.get('port_atr_bdr', ''),
                'delta_port_atr': row.get('delta_port_atr', ''),
                'lag_atr': row.get('lag_atr', ''),
                'lag_atr_bdr': row.get('lag_atr_bdr', ''),
                'delta_lag_atr': row.get('delta_lag_atr', ''),
                'loopback0_atr': row.get('loopback0_atr', ''),
                'lo_ip': row.get('lo_ip', ''),
                'delta_loopback0_atr': row.get('delta_loopback0_atr', ''),
                'loopback1_atr': row.get('loopback1_atr', ''),
                'l1_ip': row.get('l1_ip', ''),
                'delta_loopback1_atr': row.get('delta_loopback1_atr', ''),
                'ip_interco_atr': row.get('ip_interco_atr', ''),
                'ip_interco_bdr': row.get('ip_interco_bdr', ''),
                'delta_ip_interco_atr': row.get('delta_ip_interco_atr', ''),
                'vlan_interco': row.get('vlan_interco', ''),
                'vlan_interco_bdr': row.get('vlan_interco_bdr', ''),
                'delta_vlan_interco_atr': row.get('delta_vlan_interco_atr', ''),
                'port_parent': row.get('port_parent', ''),
                'port_parent_bdr': row.get('port_parent_bdr', ''),
                'delta_port_parent': row.get('delta_port_parent', ''),
                'lag_parent': row.get('lag_parent', ''),
                'lag_parent_bdr': row.get('lag_parent_bdr', ''),
                'delta_lag_parent': row.get('delta_lag_parent', ''),
                'ip_interco_parent': row.get('ip_interco_parent', ''),
                'ip_interco_parent_bdr': row.get('ip_interco_parent_bdr', ''),
                'delta_ip_interco_parent': row.get('delta_ip_interco_parent', ''),
                'vlan_interco_parent': row.get('vlan_interco_parent', ''),
                'vlan_interco_parent_bdr': row.get('vlan_interco_parent_bdr', ''),
                'delta_vlan_interco_parent': row.get('delta_vlan_interco_parent', ''),
                'port': row.get('port', ''),
                'lag': row.get('lag', ''),
                'vlan': row.get('vlan', ''),
                'addresse_ip': row.get('addresse_ip', ''),
                'etat_parent': row.get('etat_parent', ''),
                'etat_des_couples': row.get('etat_des_couples', ''),
                'etat_final': row.get('etat_final', ''),
                'etat_gcr': row.get('etat_gcr', ''),
                'service': row.get('service', ''),
                'responsable': row.get('responsable', ''),

                # If insertion_date is in the file and properly formatted,
                # you can parse it directly. If it's not guaranteed, you might
                # need more robust parsing or set it to None.
                'insertion_date': row.get('insertion_date', None),
            }

            AuditATR.objects.create(**data)

        return {"message": "Excel data processed and saved successfully!"}

    def process_csv_atr(self, file):
        df = pd.read_csv(file)

        # Iterate over each row in the DataFrame
        for _, row in df.iterrows():
            data = {
                'key': row.get('Key', ''),
                'atr': row.get('ATR', ''),
                'connexion': row.get('connexion', ''),
                'equipment_parent': row.get('equipment_parent', ''),
                'equipement_parent_bdr': row.get('equipement_parent_bdr', ''),
                'delta_equipement_parent': row.get('delta_equipement_parent', ''),
                'port_atr': row.get('port_atr', ''),
                'port_atr_bdr': row.get('port_atr_bdr', ''),
                'delta_port_atr': row.get('delta_port_atr', ''),
                'lag_atr': row.get('lag_atr', ''),
                'lag_atr_bdr': row.get('lag_atr_bdr', ''),
                'delta_lag_atr': row.get('delta_lag_atr', ''),
                'loopback0_atr': row.get('loopback0_atr', ''),
                'lo_ip': row.get('lo_ip', ''),
                'delta_loopback0_atr': row.get('delta_loopback0_atr', ''),
                'loopback1_atr': row.get('loopback1_atr', ''),
                'l1_ip': row.get('l1_ip', ''),
                'delta_loopback1_atr': row.get('delta_loopback1_atr', ''),
                'ip_interco_atr': row.get('ip_interco_atr', ''),
                'ip_interco_bdr': row.get('ip_interco_bdr', ''),
                'delta_ip_interco_atr': row.get('delta_ip_interco_atr', ''),
                'vlan_interco': row.get('vlan_interco', ''),
                'vlan_interco_bdr': row.get('vlan_interco_bdr', ''),
                'delta_vlan_interco_atr': row.get('delta_vlan_interco_atr', ''),
                'port_parent': row.get('port_parent', ''),
                'port_parent_bdr': row.get('port_parent_bdr', ''),
                'delta_port_parent': row.get('delta_port_parent', ''),
                'lag_parent': row.get('lag_parent', ''),
                'lag_parent_bdr': row.get('lag_parent_bdr', ''),
                'delta_lag_parent': row.get('delta_lag_parent', ''),
                'ip_interco_parent': row.get('ip_interco_parent', ''),
                'ip_interco_parent_bdr': row.get('ip_interco_parent_bdr', ''),
                'delta_ip_interco_parent': row.get('delta_ip_interco_parent', ''),
                'vlan_interco_parent': row.get('vlan_interco_parent', ''),
                'vlan_interco_parent_bdr': row.get('vlan_interco_parent_bdr', ''),
                'delta_vlan_interco_parent': row.get('delta_vlan_interco_parent', ''),
                'port': row.get('port', ''),
                'lag': row.get('lag', ''),
                'vlan': row.get('vlan', ''),
                'addresse_ip': row.get('addresse_ip', ''),
                'etat_parent': row.get('etat_parent', ''),
                'etat_des_couples': row.get('etat_des_couples', ''),
                'etat_final': row.get('etat_final', ''),
                'etat_gcr': row.get('etat_gcr', ''),
                'service': row.get('service', ''),
                'responsable': row.get('responsable', ''),
                # Same note about insertion_date parsing applies here
                'insertion_date': row.get('insertion_date', None),
            }

            AuditATR.objects.create(**data)

        return {"message": "CSV data processed and saved successfully!"}

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        if not file:
            return Response({"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Detect file type based on extension
            if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
                result = self.process_excel_atr(file)
            elif file.name.endswith('.csv'):
                result = self.process_csv_atr(file)
            else:
                return Response(
                    {"error": "Unsupported file format. Please upload .xlsx, .xls, or .csv"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            return Response(result, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
class AuditAtrRadio(APIView):
    def get():
        return None

class AuditAtrImes(APIView):
    def getLatestDate(self):
        latest_record = (
            AuditATR.objects.exclude(insertion_date__isnull=True)
            .order_by('-insertion_date')
            .values('insertion_date')
            .first()
        )
        # If no records exist, return an empty response or handle as needed
        if not latest_record:
            # Return None (or a Python date object, or an empty string)
            return None
        
        latest_month = latest_record['insertion_date']

        return latest_month

    def tauxdecoherence(self):
        # Determine the latest month
        latest_month = self.getLatestDate()

        # Count the total number of rows with the latest month
        total_records = AuditATR.objects.filter(
            insertion_date=latest_month
        ).count()

        # Count the number of rows with 'CSG-FH-CTR' in the Architecture column for the latest month
        atr_count = AuditATR.objects.filter(
            insertion_date=latest_month
        ).values('atr').distinct().count()
        total_atr_count = AuditATR.objects.values('atr').distinct().count()

        # Calculate the percentage
        percentage = (atr_count / total_atr_count) * 100 if total_records > 0 else 0.0
        percentage = round(percentage)
        result_data = {
            'percentage': percentage,
            'total_records': total_records,
            'atr_count': atr_count,
        }

        return result_data

    def repartitionAtr(self):
        # Get all records with value 'Site avec au moins une incohérence' in ETAT_CSG
        latest_date = self.getLatestDate()
        records = AuditATR.objects.filter(etat_final='Site avec au moins une incoherence', insertion_date=latest_date)
        
        # Columns to check for NOK values
        columns_to_check = [
            'delta_vlan_interco_parent', 'delta_ip_interco_parent', 'delta_lag_parent', 
            'delta_port_parent', 'delta_vlan_interco_atr','delta_ip_interco_atr', 'delta_loopback1_atr', 'delta_loopback0_atr',
            'delta_lag_atr', 'delta_port_atr','delta_equipement_parent'
        ]

        # Initialize counts
        one_nok_count = 0
        two_nok_count = 0
        three_or_more_nok_count = 0

        # Iterate over each record
        for record in records:
            nok_count = sum(1 for column in columns_to_check if getattr(record, column) == 'NOK')
            
            if nok_count == 1:
                one_nok_count += 1
            elif nok_count == 2:
                two_nok_count += 1
            elif nok_count >= 3:
                three_or_more_nok_count += 1

        # Return the counts as JSON response
        data = {
            'one_nok_count': one_nok_count,
            'two_nok_count': two_nok_count,
            'three_or_more_nok_count': three_or_more_nok_count,
        }

        return data
    def responsable(self):
        # Dictionary to store the count of each entity
        responsable_counts = defaultdict(int)

        # Get the latest date across all records
        latest_date = self.getLatestDate()

        if not latest_date:
            return {}  # Return empty if there's no data

        # Filter records for the latest date and where etat_final = "Site avec au moins une incohérence"
        filtered_data = AuditATR.objects.filter(
            insertion_date=latest_date,
            etat_final="Site avec au moins une incoherence"
        )

        # Aggregate the count of each unique value in the 'responsable' column
        usine_prod_counts = filtered_data.values('responsable').annotate(count=Count('responsable'))

        # Process each entry to correctly distribute counts
        for item in usine_prod_counts:
            responsables = item['responsable']
            count = item['count']

            # Ensure responsible field is not empty or 'nan'
            if responsables and str(responsables).strip().lower() not in ["nan", ""]:
                # Split and distribute the count to each entity
                for entity in responsables.split(','):
                    responsable_counts[entity.strip()] += count

        # Sort by count in ascending order
        sorted_responsable_counts = dict(
            sorted(responsable_counts.items(), key=lambda x: x[1])
        )

        return sorted_responsable_counts

    def calculate_nok_percentage(self):
        latest_month = self.getLatestDate()

        # Mapping column names to user-friendly labels
        column_labels = {
            'addresse_ip': 'Addresse IP',
            'vlan': 'Vlan',
            'lag': 'LAG',
            'port': 'Port'
        }

        # Initialize counts
        nok_counts = {column: 0 for column in column_labels}
        total_nok_count = 0

        # Fetch records for the given architecture
        records = AuditATR.objects.filter(insertion_date=latest_month)

        # Iterate through each record to calculate NOK counts
        for record in records:
            # Check and count NOKs for standard columns
            for column in column_labels:
                if getattr(record, column, 'NOK') == 'NOK':
                    nok_counts[column] += 1
                    total_nok_count += 1

        # If there are no NOKs, return zero for all fields
        if total_nok_count == 0:
            return {label: 0 for label in column_labels.values()}

        # Calculate percentages for each column
        nok_percentages = {}
        for column, count in nok_counts.items():
            label = column_labels[column]
            nok_percentages[label] = round(count)

        return nok_percentages

    def evolutionDeCoherance(self):
        # 1. Fetch the latest insertion_date for this architecture
        latest_record = (
            AuditATR.objects
            .order_by('-insertion_date')
            .values('insertion_date')
            .first()
        )

        if not latest_record:
            return []

        latest_insertion_date = latest_record['insertion_date']
        
        # 2. Determine the start/end of the latest month
        latest_month_start = latest_insertion_date.replace(day=1)
        latest_month_end = latest_month_start + relativedelta(months=1, days=-1)
        
        # 3. Determine the start/end of the previous month
        previous_month_end = latest_month_start - timedelta(days=1)
        previous_month_start = previous_month_end.replace(day=1)

        # 4. Query for the LATEST month
        grouped_data_latest_month = (
            AuditATR.objects.filter(
                insertion_date__range=(latest_month_start, latest_month_end)
            )
            .values('insertion_date')
            .annotate(
                total_records=Count('key'),
                sites_conformes=Count(
                    Case(When(etat_final='Site Conforme', then=Value(1)))
                )
            )
        )
    
        # 5. Query for the PREVIOUS month
        grouped_data_previous_month = (
            AuditATR.objects.filter(
                insertion_date__range=(previous_month_start, previous_month_end)
            )
            .values('insertion_date')
            .annotate(
                total_records=Count('key'),
                sites_conformes=Count(
                    Case(When(etat_final='Site Conforme', then=Value(1)))
                )
            )
        )

        # 6. Calculate percentages and differences
        result_data = []

        # We zip these two QuerySets so we can compare them month-by-month
        for item_latest, item_previous in zip(grouped_data_latest_month, grouped_data_previous_month):
            total_records_latest = item_latest['total_records']
            sites_conformes_latest = item_latest['sites_conformes']
            percentage_latest = (
                (sites_conformes_latest / total_records_latest) * 100
                if total_records_latest > 0
                else 0.0
            )
            percentage_latest = round(percentage_latest)

            total_records_previous = item_previous['total_records']
            sites_conformes_previous = item_previous['sites_conformes']
            percentage_previous = (
                (sites_conformes_previous / total_records_previous) * 100
                if total_records_previous > 0
                else 0.0
            )

            difference = percentage_latest - percentage_previous
            difference = round(difference)

            result_data.append({
                # Convert insertion_date to string for serialization
                'date': item_latest['insertion_date'].strftime("%Y-%m-%d"),
                'difference': abs(difference),
                'positive': difference >= 0,
                'value': f"{round(percentage_latest)}%", 
                'category': self.calculate_category(percentage_latest),
            })

        return result_data

    def calculate_category(self,percentage):
        if percentage < 60:
            return 'low'
        elif 60 <= percentage <= 90:
            return 'medium'
        else:
            return 'high'
    
    def evolution(self):
        filtered_data = AuditATR.objects.all()
        
        grouped_data = (
            filtered_data
            .annotate(month=TruncMonth('insertion_date'))
            .values('month')
            .annotate(
                sites_conformes=Count(Case(When(etat_final='Site Conforme', then=1))),
                sites_avec_incoherence=Count(Case(When(etat_final='Site avec au moins une incoherence', then=1)))
            )
            .order_by('-month')[:11]
        )

        # Format the 'month' field properly
        formatted_results = []
        for row in grouped_data:
            # original_month = row['month']  # e.g., datetime.date(2025, 2, 1)
            # row['month'] = original_month.strftime('%b-%Y')  # Convert "2025-02-01" ? "Feb-2025"
            formatted_results.append(row)

        return formatted_results

    def get(self, request, format=None):
        # Extract the architecture from the query parameters (if provided)
        try:
            # Fetch data from each method
            latest_date = self.getLatestDate()

            if not latest_date:
                # If there is no record, handle however you like
                return Response(
                    {
                        "latest_date": None,
                        "message": f"No records found for ATR"
                    },
                    status=status.HTTP_200_OK
                )

            taux_de_coherence = self.tauxdecoherence()
            responsable = self.responsable()
            nok_percentage = self.calculate_nok_percentage()
            evolution_de_coherance = self.evolutionDeCoherance()
            evolution_data = self.evolution()

            # Prepare the final response
            response_data = {
                'latest_date': latest_date,
                'taux_de_coherence': taux_de_coherence,
                'responsable': responsable,
                'nok_percentage': nok_percentage,
                'evolution_de_coherance': evolution_de_coherance,
                'evolution': list(evolution_data),  # Convert queryset to list for serialization
            }

            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class CustomAuditView(APIView):
    def post(self, request, *args, **kwargs):
        """
        Handle the file upload and save the audit record.
        """
        # Ensure that title and image are part of the request
        title = request.data.get("title")
        image = request.FILES.get("image")
        audit_type = request.data.get("audit_type")

        if not title or not image:
            return Response({"error": "Title and image are required."}, status=status.HTTP_400_BAD_REQUEST)

        # Create a new audit record
        try:
            custom_audit = CustomAudit.objects.create(title=title, image=image,audit_type=audit_type)
            custom_audit.save()

            # Optionally, return the created object's data
            return Response(CustomAuditSerializer(custom_audit).data, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def get(self, request, *args, **kwargs):
        """
        Return the list of audits filtered by audit_type.
        """
        audit_type = request.query_params.get("audit_type")

        if audit_type:
            audits = CustomAudit.objects.filter(audit_type=audit_type)
        else:
            audits = CustomAudit.objects.all()

        # Serialize the queryset
        serializer = CustomAuditSerializer(audits, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
class EditCustomAuditView(APIView):
    def post(self, request, pk, *args, **kwargs):
        """
        Edit (update) an existing audit record.
        """
        try:
            custom_audit = CustomAudit.objects.get(pk=pk)
        except CustomAudit.DoesNotExist:
            return Response({"error": "Audit not found."},
                            status=status.HTTP_404_NOT_FOUND)

        title = request.data.get("title")
        image = request.FILES.get("image")  # Optional update
        audit_type = request.data.get("audit_type")  # Optional update

        if not title:
            return Response({"error": "Title is required."},
                            status=status.HTTP_400_BAD_REQUEST)

        custom_audit.title = title
        if image:
            custom_audit.image = image
        if audit_type:
            custom_audit.audit_type = audit_type

        try:
            custom_audit.save()
            serializer = CustomAuditSerializer(custom_audit)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class DeleteCustomAuditView(APIView):
    def post(self, request, pk, *args, **kwargs):
        try:
            custom_audit = CustomAudit.objects.get(pk=pk)
        except CustomAudit.DoesNotExist:
            return Response({"error": "Audit not found."}, status=status.HTTP_404_NOT_FOUND)
        try:
            custom_audit.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)








###### Audit Doublons IP ######
class AuditDoublonsIPUploadView(APIView):

    def process_excel_doublons_ip(self, file):
        df = pd.read_excel(file)

        # Iterate over each row in the DataFrame
        for _, row in df.iterrows():
            data = {
                "doublons_adresse_ip": row.get("Doublons Adresse IP"),
                "add_ip_paire": row.get("Add IP Paire"),
                "add_ip_impaire": row.get("Add IP Impaire"),
                "equipment_radio_dans_bdr_trans": row.get("Equipement Radio BDR Trans"),
                "equipment_trans_dans_bdr_trans": row.get("Equipement Trans BDR Trans"),
                "equipment_radio_dans_bde_trans": row.get("Equipement Radio BDE Trans"),
                "equipment_trans_dans_bde_trans": row.get("Equipement Trans BDE Trans"),
                "equipment_radio_dans_bde_radio": row.get("Equipement Radio BDE Radio"),
                "port_dans_bde_trans": row.get("Port dans BDE Trans"),
                "existe_dans_bdr_trans": row.get("Existe dans BDR Trans"),
                "doublons_dans_bde_trans": row.get("Doublons BDE Trans"),
                "doublons_dans_bde_radio": row.get("Doublons BDE Radio"),
                "insertion_date": row.get("insertion_date")
            }

            AuditDoublonsIP.objects.create(**data)

        return {"message": "Excel data processed and saved successfully!"}

    def process_csv_doublons_ip(self, file):
        df = pd.read_csv(file)

        # Iterate over each row in the DataFrame
        for _, row in df.iterrows():
            data = {
                "doublons_adresse_ip": row.get("Doublons Adresse IP"),
                "add_ip_paire": row.get("Add IP Paire"),
                "add_ip_impaire": row.get("Add IP Impaire"),
                "equipment_radio_dans_bdr_trans": row.get("Equipement Radio BDR Trans"),
                "equipment_trans_dans_bdr_trans": row.get("Equipement Trans BDR Trans"),
                "equipment_radio_dans_bde_trans": row.get("Equipement Radio BDE Trans"),
                "equipment_trans_dans_bde_trans": row.get("Equipement Trans BDE Trans"),
                "equipment_radio_dans_bde_radio": row.get("Equipement Radio BDE Radio"),
                "port_dans_bde_trans": row.get("Port dans BDE Trans"),
                "existe_dans_bdr_trans": row.get("Existe dans BDR Trans"),
                "doublons_dans_bde_trans": row.get("Doublons BDE Trans"),
                "doublons_dans_bde_radio": row.get("Doublons BDE Radio"),
                "insertion_date": row.get("insertion_date")
            }

            AuditDoublonsIP.objects.create(**data)

        return {"message": "CSV data processed and saved successfully!"}

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        if not file:
            return Response({"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Detect file type based on extension
            if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
                result = self.process_excel_doublons_ip(file)
            elif file.name.endswith('.csv'):
                result = self.process_csv_doublons_ip(file)
            else:
                return Response(
                    {"error": "Unsupported file format. Please upload .xlsx, .xls, or .csv"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            return Response(result, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)







# class AuditDoublonsIPView(APIView):
#     def getLatestDate(self):
#         latest_record = (
#             AuditDoublonsIP.objects.exclude(insertion_date__isnull=True)
#             .order_by('-insertion_date')
#             .values('insertion_date')
#             .first()
#         )
#         # If no records exist, return an empty response or handle as needed
#         if not latest_record:
#             # Return None (or a Python date object, or an empty string)
#             return None
        
#         latest_month = latest_record['insertion_date']

#         return latest_month

#     def tauxdecoherence(self):
#         latest_month = self.getLatestDate()
#         if not latest_month:
#             return {'percentage': 0, 'total_records': 0, 'doublons_ip_count': 0}

#         total_records = AuditDoublonsIP.objects.filter(insertion_date=latest_month).count()
#         doublons_ip_count = AuditDoublonsIP.objects.filter(
#             insertion_date=latest_month
#         ).values('doublons_adresse_ip').distinct().count()

#         percentage = (doublons_ip_count / total_records) * 100 if total_records > 0 else 0.0
#         percentage = round(percentage)

#         return {
#             'percentage': percentage,
#             'total_records': total_records,
#             'doublons_ip_count': doublons_ip_count,
#         }


#     # def repartitionAtr(self):
#     #     # Get all records with value 'Site avec au moins une incohérence' in ETAT_CSG
#     #     latest_date = self.getLatestDate()
#     #     records = AuditDoublonsIP.objects.filter(etat_final='Site avec au moins une incoherence', insertion_date=latest_date)
        
#     #     # Columns to check for NOK values
#     #     columns_to_check = [
#     #         'delta_vlan_interco_parent', 'delta_ip_interco_parent', 'delta_lag_parent', 
#     #         'delta_port_parent', 'delta_vlan_interco_atr','delta_ip_interco_atr', 'delta_loopback1_atr', 'delta_loopback0_atr',
#     #         'delta_lag_atr', 'delta_port_atr','delta_equipement_parent'
#     #     ]

#     #     # Initialize counts
#     #     one_nok_count = 0
#     #     two_nok_count = 0
#     #     three_or_more_nok_count = 0

#     #     # Iterate over each record
#     #     for record in records:
#     #         nok_count = sum(1 for column in columns_to_check if getattr(record, column) == 'NOK')
            
#     #         if nok_count == 1:
#     #             one_nok_count += 1
#     #         elif nok_count == 2:
#     #             two_nok_count += 1
#     #         elif nok_count >= 3:
#     #             three_or_more_nok_count += 1

#     #     # Return the counts as JSON response
#     #     data = {
#     #         'one_nok_count': one_nok_count,
#     #         'two_nok_count': two_nok_count,
#     #         'three_or_more_nok_count': three_or_more_nok_count,
#     #     }

#     #     return data
#     # def responsable(self):
#     #     # Dictionary to store the count of each entity
#     #     responsable_counts = defaultdict(int)

#     #     # Get the latest date across all records
#     #     latest_date = self.getLatestDate()

#     #     if not latest_date:
#     #         return {}  # Return empty if there's no data

#     #     # Filter records for the latest date and where etat_final = "Site avec au moins une incohérence"
#     #     filtered_data = AuditDoublonsIP.objects.filter(
#     #         insertion_date=latest_date,
#     #         etat_final="Site avec au moins une incoherence"
#     #     )

#     #     # Aggregate the count of each unique value in the 'responsable' column
#     #     usine_prod_counts = filtered_data.values('responsable').annotate(count=Count('responsable'))

#     #     # Process each entry to correctly distribute counts
#     #     for item in usine_prod_counts:
#     #         responsables = item['responsable']
#     #         count = item['count']

#     #         # Ensure responsible field is not empty or 'nan'
#     #         if responsables and str(responsables).strip().lower() not in ["nan", ""]:
#     #             # Split and distribute the count to each entity
#     #             for entity in responsables.split(','):
#     #                 responsable_counts[entity.strip()] += count

#     #     # Sort by count in ascending order
#     #     sorted_responsable_counts = dict(
#     #         sorted(responsable_counts.items(), key=lambda x: x[1])
#     #     )

#     #     return sorted_responsable_counts

#     # def calculate_nok_percentage(self):
#     #     latest_month = self.getLatestDate()

#     #     # Mapping column names to user-friendly labels
#     #     column_labels = {
#     #         'addresse_ip': 'Addresse IP',
#     #         'vlan': 'Vlan',
#     #         'lag': 'LAG',
#     #         'port': 'Port'
#     #     }

#     #     # Initialize counts
#     #     nok_counts = {column: 0 for column in column_labels}
#     #     total_nok_count = 0

#     #     # Fetch records for the given architecture
#     #     records = AuditDoublonsIP.objects.filter(insertion_date=latest_month)

#     #     # Iterate through each record to calculate NOK counts
#     #     for record in records:
#     #         # Check and count NOKs for standard columns
#     #         for column in column_labels:
#     #             if getattr(record, column, 'NOK') == 'NOK':
#     #                 nok_counts[column] += 1
#     #                 total_nok_count += 1

#     #     # If there are no NOKs, return zero for all fields
#     #     if total_nok_count == 0:
#     #         return {label: 0 for label in column_labels.values()}

#     #     # Calculate percentages for each column
#     #     nok_percentages = {}
#     #     for column, count in nok_counts.items():
#     #         label = column_labels[column]
#     #         nok_percentages[label] = round(count)

#     #     return nok_percentages

#     # def evolutionDeCoherance(self):
#     #     # 1. Fetch the latest insertion_date for this architecture
#     #     latest_record = (
#     #         AuditDoublonsIP.objects
#     #         .order_by('-insertion_date')
#     #         .values('insertion_date')
#     #         .first()
#     #     )

#     #     if not latest_record:
#     #         return []

#     #     latest_insertion_date = latest_record['insertion_date']
        
#     #     # 2. Determine the start/end of the latest month
#     #     latest_month_start = latest_insertion_date.replace(day=1)
#     #     latest_month_end = latest_month_start + relativedelta(months=1, days=-1)
        
#     #     # 3. Determine the start/end of the previous month
#     #     previous_month_end = latest_month_start - timedelta(days=1)
#     #     previous_month_start = previous_month_end.replace(day=1)

#     #     # 4. Query for the LATEST month
#     #     grouped_data_latest_month = (
#     #         AuditDoublonsIP.objects.filter(
#     #             insertion_date__range=(latest_month_start, latest_month_end)
#     #         )
#     #         .values('insertion_date')
#     #         .annotate(
#     #             total_records=Count('key'),
#     #             sites_conformes=Count(
#     #                 Case(When(etat_final='Site Conforme', then=Value(1)))
#     #             )
#     #         )
#     #     )
    
#     #     # 5. Query for the PREVIOUS month
#     #     grouped_data_previous_month = (
#     #         AuditDoublonsIP.objects.filter(
#     #             insertion_date__range=(previous_month_start, previous_month_end)
#     #         )
#     #         .values('insertion_date')
#     #         .annotate(
#     #             total_records=Count('key'),
#     #             sites_conformes=Count(
#     #                 Case(When(etat_final='Site Conforme', then=Value(1)))
#     #             )
#     #         )
#     #     )

#     #     # 6. Calculate percentages and differences
#     #     result_data = []

#     #     # We zip these two QuerySets so we can compare them month-by-month
#     #     for item_latest, item_previous in zip(grouped_data_latest_month, grouped_data_previous_month):
#     #         total_records_latest = item_latest['total_records']
#     #         sites_conformes_latest = item_latest['sites_conformes']
#     #         percentage_latest = (
#     #             (sites_conformes_latest / total_records_latest) * 100
#     #             if total_records_latest > 0
#     #             else 0.0
#     #         )
#     #         percentage_latest = round(percentage_latest)

#     #         total_records_previous = item_previous['total_records']
#     #         sites_conformes_previous = item_previous['sites_conformes']
#     #         percentage_previous = (
#     #             (sites_conformes_previous / total_records_previous) * 100
#     #             if total_records_previous > 0
#     #             else 0.0
#     #         )

#     #         difference = percentage_latest - percentage_previous
#     #         difference = round(difference)

#     #         result_data.append({
#     #             # Convert insertion_date to string for serialization
#     #             'date': item_latest['insertion_date'].strftime("%Y-%m-%d"),
#     #             'difference': abs(difference),
#     #             'positive': difference >= 0,
#     #             'value': f"{round(percentage_latest)}%", 
#     #             'category': self.calculate_category(percentage_latest),
#     #         })

#     #     return result_data

#     # def calculate_category(self,percentage):
#     #     if percentage < 60:
#     #         return 'low'
#     #     elif 60 <= percentage <= 90:
#     #         return 'medium'
#     #     else:
#     #         return 'high'
    
#     # def evolution(self):
#     #     filtered_data = AuditDoublonsIP.objects.all()
        
#     #     grouped_data = (
#     #         filtered_data
#     #         .annotate(month=TruncMonth('insertion_date'))
#     #         .values('month')
#     #         .annotate(
#     #             sites_conformes=Count(Case(When(etat_final='Site Conforme', then=1))),
#     #             sites_avec_incoherence=Count(Case(When(etat_final='Site avec au moins une incoherence', then=1)))
#     #         )
#     #         .order_by('-month')[:11]
#     #     )

#     #     # Format the 'month' field properly
#     #     formatted_results = []
#     #     for row in grouped_data:
#     #         # original_month = row['month']  # e.g., datetime.date(2025, 2, 1)
#     #         # row['month'] = original_month.strftime('%b-%Y')  # Convert "2025-02-01" ? "Feb-2025"
#     #         formatted_results.append(row)

#     #     return formatted_results

#     def getLatestDate(self, architecture):
#         latest_record = (
#             AuditDoublonsIP.objects.filter(architecture=architecture).order_by('-insertion_date').values('insertion_date').first()
#         )
#         # If no records exist, return an empty response or handle as needed
#         if not latest_record:
#             return None
        
#         latest_month = latest_record['insertion_date']
#         return latest_month




#     def get(self, request, format=None):
#         # Extract the architecture from the query parameters (if provided)
#         try:
#             # Fetch data from each method
#             latest_date = self.getLatestDate()

#             if not latest_date:
#                 return Response(
#                     {
#                         "latest_date": None,
#                         "message": f"No records found for IP"
#                     },
#                     status=status.HTTP_200_OK
#                 )
#             taux_de_coherence = self.tauxdecoherence(architecture)
#             # responsable = self.responsable()
#             # nok_percentage = self.calculate_nok_percentage()
#             # evolution_de_coherance = self.evolutionDeCoherance()
#             # evolution_data = self.evolution()

#             # Prepare the final response
#             response_data = {
#                 'latest_date': latest_date,
#                 'taux_de_coherence': taux_de_coherence,
#                 # 'responsable': responsable,
#                 # 'nok_percentage': nok_percentage,
#                 # 'evolution_de_coherance': evolution_de_coherance,
#                 # 'evolution': list(evolution_data),  # Convert queryset to list for serialization
#             }

#             return Response(response_data, status=status.HTTP_200_OK)

#         except Exception as e:
#             return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class AuditDoublonsIPView(APIView):
    """
    API view to return the latest audit date and coherence rate.
    Handles empty tables safely.
    """

    def get_latest_date(self, architecture=None):
        """
        Get the latest insertion date, optionally filtered by architecture.
        Returns None if no records exist.
        """
        qs = AuditDoublonsIP.objects.all()
        if architecture:
            qs = qs.filter(architecture=architecture)

        latest_record = qs.exclude(insertion_date__isnull=True).order_by('-insertion_date').values('insertion_date').first()
        if not latest_record:
            return None

        return latest_record['insertion_date']


    def taux_de_coherence(self, architecture=None):
        """
        Calculate taux de coherence:
        (rows with 'Unique dans BDE TRANS' in K + rows with 'pas de doublons' in L)
        divided by total rows.
        """
        latest_date = self.get_latest_date(architecture)
        if not latest_date:
            return {'percentage': 0, 'total_records': 0, 'bde_trans_numerator': 0}

        # Base queryset
        qs = AuditDoublonsIP.objects.filter(insertion_date=latest_date)
        if architecture:
            qs = qs.filter(architecture=architecture)

        total_records = qs.count()

        # Numerator: rows that match either condition
        numerator = qs.filter(
            Q(doublons_bde_trans="Unique dans BDE TRANS") |
            Q(doublons_bde_radio="pas de doublons")
        ).count()

        percentage = round((numerator / total_records) * 100) if total_records else 0

        return {
            "percentage": percentage,
            "total_records": total_records,
            "doublons_ip_count": bde_trans_numerator,  # alias
        }



    def get(self, request, format=None):
        try:
            architecture = request.query_params.get('architecture', None)

            latest_date = self.get_latest_date(architecture)
            if not latest_date:
                return Response({
                    "latest_date": None,
                    "message": "No records found for IP"
                }, status=status.HTTP_200_OK)

            taux_de_coherence = self.taux_de_coherence(architecture)

            response_data = {
                'latest_date': latest_date,
                'taux_de_coherence': taux_de_coherence,
            }

            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            # Log the exception in production for debugging
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
